# coding=utf-8
"""
股票池选股器 V2.1 - 改进版
改进内容：
1. 自动查找配置文件
2. 内置默认配置（无需外部文件也能运行）
3. 更友好的错误提示
4. 简化初始化流程
"""
from __future__ import print_function, absolute_import, unicode_literals
from gm.api import *
import datetime
import numpy as np
import pandas as pd
import logging
import json
import math
import os
from collections import OrderedDict

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler('stock_selector_v2.log', encoding='utf-8'), logging.StreamHandler()]
)
logger = logging.getLogger(__name__)


class StockSelectorV2:
    """
    股票池选股器 V2.1 - 增强版
    新增特性：
    1. 新增更多因子（股息率、贝塔、反转、量比）
    2. 动态权重调整（根据市场环境）
    3. 行业中性化处理
    4. 多时点选股支持
    5. 自动查找配置文件
    """

    def __init__(self, config_file=None):
        """
        初始化选股器
        
        参数:
            config_file: 配置文件路径（可选，不提供则使用内置默认配置）
        """
        self.load_config(config_file)
        self.stock_pool = []
        self.stock_analysis = {}
        self.market_environment = 'neutral'

    def _find_config_file(self):
        """自动查找配置文件"""
        possible_paths = [
            'stock_selector_v2_config.json',
            './stock_selector_v2_config.json',
            '../stock_selector_v2_config.json',
        ]
        
        # 也检查当前工作目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        if script_dir:
            possible_paths.append(os.path.join(script_dir, 'stock_selector_v2_config.json'))
        
        for path in possible_paths:
            if os.path.exists(path):
                logger.info(f'找到配置文件: {path}')
                return path
        
        return None

    def load_config(self, config_file=None):
        """加载配置"""
        # 情况1：用户指定了配置文件
        if config_file and os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    self.config = json.load(f)
                logger.info(f'配置加载成功: {config_file}')
                self._validate_token()
                return
            except Exception as e:
                logger.warning(f'加载配置文件失败: {e}')
        
        # 情况2：自动查找配置文件
        auto_config = self._find_config_file()
        if auto_config:
            try:
                with open(auto_config, 'r', encoding='utf-8') as f:
                    self.config = json.load(f)
                logger.info(f'配置加载成功: {auto_config}')
                self._validate_token()
                return
            except Exception as e:
                logger.warning(f'加载配置文件失败: {e}')
        
        # 情况3：使用内置默认配置
        logger.info('使用内置默认配置')
        self.config = self._get_default_config()
        self._validate_token()

    def _validate_token(self):
        """验证Token配置"""
        token = self.config.get('backtest', {}).get('token', '')
        if token == '{{token}}' or not token or len(token) < 10:
            logger.warning('⚠️ 注意：Token未配置或可能无效')
            logger.warning('   请在掘金终端中运行，或配置正确的Token')

    def _get_default_config(self):
        """获取内置默认配置"""
        return {
            "strategy": {
                "name": "多因子选股策略V2.1",
                "lookback_days_short": 10,
                "lookback_days_mid": 20,
                "lookback_days_long": 60,
                "holding_num": 25,
                "min_market_cap": 50,
                "max_pe": 120,
                "max_pb": 12,
                "enable_industry_neutral": True
            },
            "indices": {
                "SHSE.000300": "沪深300",
                "SHSE.000905": "中证500",
                "SHSE.000016": "上证50",
                "SZSE.399005": "中小板指"
            },
            "factors": {
                "momentum_short_weight": 0.12,
                "momentum_mid_weight": 0.15,
                "momentum_long_weight": 0.08,
                "valuation_pe_weight": 0.08,
                "valuation_pb_weight": 0.06,
                "valuation_ps_weight": 0.04,
                "dividend_yield_weight": 0.06,
                "volatility_weight": 0.08,
                "quality_roe_weight": 0.10,
                "growth_weight": 0.08,
                "reversal_weight": 0.05,
                "beta_weight": 0.04,
                "volume_ratio_weight": 0.04,
                "size_weight": 0.02
            },
            "market_adjust": {
                "bull": {"momentum": 1.4, "value": 0.9, "quality": 0.9, "volatility": 0.8},
                "bear": {"momentum": 0.7, "value": 1.3, "quality": 1.4, "volatility": 1.4},
                "volatile": {"momentum": 0.85, "value": 1.1, "quality": 1.2, "volatility": 1.3},
                "neutral": {"momentum": 1.0, "value": 1.0, "quality": 1.0, "volatility": 1.0}
            },
            "backtest": {
                "token": ""
            }
        }

    def judge_market_environment(self, date_str):
        """判断市场环境"""
        try:
            data = history_n(
                symbol='SHSE.000300',
                frequency='1d',
                count=30,
                fields='close,volume',
                fill_missing='Last',
                adjust=ADJUST_PREV,
                end_time=date_str,
                df=True
            )

            if data is None or len(data) < 15:
                logger.warning('市场数据不足，默认使用neutral环境')
                return 'neutral'

            returns = data['close'].pct_change().dropna()
            avg_return = returns.mean()
            volatility = returns.std()
            volume_trend = (data['volume'].iloc[-5:].mean() / data['volume'].iloc[-20:-5].mean()) - 1

            if avg_return > 0.003 and volume_trend > 0.1:
                return 'bull'
            elif avg_return < -0.003:
                return 'bear'
            elif volatility > 0.028:
                return 'volatile'
            else:
                return 'neutral'
        except Exception as e:
            logger.error(f'判断市场环境失败：{e}')
            return 'neutral'

    def _calculate_atr(self, high, low, close, period=14):
        """计算ATR"""
        tr1 = high[1:] - low[1:]
        tr2 = np.abs(high[1:] - close[:-1])
        tr3 = np.abs(low[1:] - close[:-1])
        tr = np.maximum(np.maximum(tr1, tr2), tr3)
        atr = np.zeros(len(close))
        if len(tr) >= period:
            atr[period] = np.mean(tr[:period])
            for i in range(period + 1, len(close)):
                atr[i] = (atr[i-1] * (period - 1) + tr[i-1]) / period
        return atr

    def _calculate_rsi(self, prices, period=14):
        """计算RSI"""
        deltas = np.diff(prices)
        gains = deltas.copy()
        gains[gains < 0] = 0
        losses = -deltas.copy()
        losses[losses < 0] = 0
        avg_gain = np.zeros(len(prices))
        avg_loss = np.zeros(len(prices))
        if len(gains) >= period:
            avg_gain[period] = np.mean(gains[:period])
            avg_loss[period] = np.mean(losses[:period])
            for i in range(period + 1, len(prices)):
                avg_gain[i] = (avg_gain[i-1] * (period - 1) + gains[i-1]) / period
                avg_loss[i] = (avg_loss[i-1] * (period - 1) + losses[i-1]) / period
        rs = avg_gain / (avg_loss + 1e-10)
        return 100 - (100 / (1 + rs))

    def _calculate_max_drawdown(self, prices):
        """计算最大回撤"""
        peak = prices[0]
        max_dd = 0
        for price in prices:
            if price > peak:
                peak = price
            dd = (price - peak) / peak
            if dd < max_dd:
                max_dd = dd
        return max_dd

    def _normalize(self, series):
        """Min-Max标准化"""
        min_val = series.min()
        max_val = series.max()
        return (series - min_val) / (max_val - min_val + 1e-10)

    def _calculate_beta(self, stock_prices, index_prices, period=60):
        """计算Beta系数"""
        if len(stock_prices) < period or len(index_prices) < period:
            return 1.0
        
        stock_returns = np.diff(np.log(stock_prices[-period:]))
        index_returns = np.diff(np.log(index_prices[-period:]))
        
        if len(stock_returns) < 10 or len(index_returns) < 10:
            return 1.0
        
        covariance = np.cov(stock_returns, index_returns)[0, 1]
        index_variance = np.var(index_returns)
        
        if index_variance < 1e-10:
            return 1.0
        
        beta = covariance / index_variance
        return np.clip(beta, 0.5, 1.5)

    def select_best_sector(self, end_date):
        """选择最佳风格板块"""
        self.market_environment = self.judge_market_environment(end_date)
        logger.info(f'市场环境判断: {self.market_environment}')

        scores = pd.DataFrame()
        weights = self.adjust_weights_by_market(self.config['factors'], self.market_environment)

        for index_symbol, index_name in self.config['indices'].items():
            try:
                data = history_n(
                    symbol=index_symbol,
                    frequency='1d',
                    count=60,
                    fields='close,volume,high,low',
                    fill_missing='Last',
                    adjust=ADJUST_PREV,
                    end_time=end_date,
                    df=True
                )

                if data is None or len(data) < 20:
                    continue

                prices = data['close'].values
                volumes = data['volume'].values

                momentum_short = (prices[-1] / prices[-11]) - 1 if len(prices) > 11 else 0
                momentum_mid = (prices[-1] / prices[-21]) - 1 if len(prices) > 21 else 0
                momentum_long = (prices[-1] / prices[-61]) - 1 if len(prices) > 61 else 0

                returns = np.log(prices[1:] / prices[:-1])
                volatility = returns.std() * np.sqrt(252)
                rsi = self._calculate_rsi(prices)[-1] if len(prices) >= 14 else 50
                volume_rank = np.mean(volumes[-10:] > np.roll(volumes, 10)[-10:])

                momentum_score = (
                    momentum_short * weights.get('momentum_short_weight', 0.12) +
                    momentum_mid * weights.get('momentum_mid_weight', 0.15) +
                    momentum_long * weights.get('momentum_long_weight', 0.08)
                )

                value_score = (1 - rsi / 100) * weights.get('valuation_pe_weight', 0.08)
                volatility_score = (-volatility) * weights.get('volatility_weight', 0.08)

                score = momentum_score + value_score + volatility_score + volume_rank * 0.1

                scores.loc[index_symbol, 'score'] = score
                scores.loc[index_symbol, 'name'] = index_name
                scores.loc[index_symbol, 'momentum_mid'] = momentum_mid

            except Exception as e:
                logger.warning(f'计算指数因子失败 {index_symbol}: {e}')

        if scores.empty:
            logger.warning('所有指数计算失败，默认使用沪深300')
            return 'SHSE.000300'

        best_sector = scores['score'].idxmax()
        logger.info(f'最佳板块: {best_sector} ({scores.loc[best_sector, "name"]})')

        return best_sector

    def adjust_weights_by_market(self, weights, market_env):
        """根据市场环境调整因子权重"""
        adj_weights = weights.copy()
        adjust_params = self.config.get('market_adjust', {}).get(market_env, {})

        if adjust_params.get('momentum') != 1.0:
            adj_weights['momentum_short_weight'] *= adjust_params['momentum']
            adj_weights['momentum_mid_weight'] *= adjust_params['momentum']
            adj_weights['momentum_long_weight'] *= adjust_params['momentum']

        if adjust_params.get('value') != 1.0:
            adj_weights['valuation_pe_weight'] *= adjust_params['value']
            adj_weights['valuation_pb_weight'] *= adjust_params['value']
            adj_weights['valuation_ps_weight'] *= adjust_params['value']
            adj_weights['dividend_yield_weight'] *= adjust_params['value']

        if adjust_params.get('quality') != 1.0:
            adj_weights['quality_roe_weight'] *= adjust_params['quality']
            adj_weights['growth_weight'] *= adjust_params['quality']

        if adjust_params.get('volatility') != 1.0:
            adj_weights['volatility_weight'] *= adjust_params['volatility']

        total = sum(adj_weights.values())
        return {k: v / total for k, v in adj_weights.items()}

    def get_stock_data(self, symbols, end_date):
        """获取股票数据"""
        result = []
        total = len(symbols)
        logger.info(f'开始获取 {total} 只股票数据...')

        try:
            index_data = history_n(
                symbol='SHSE.000300',
                frequency='1d',
                count=60,
                fields='close',
                fill_missing='Last',
                adjust=ADJUST_PREV,
                end_time=end_date,
                df=True
            )
            index_prices = index_data['close'].values if index_data is not None else None
        except Exception as e:
            logger.warning(f'获取指数数据失败: {e}')
            index_prices = None

        for idx, symbol in enumerate(symbols):
            try:
                price_data = history_n(
                    symbol=symbol,
                    frequency='1d',
                    count=60,
                    fields='close,volume,high,low,amount',
                    fill_missing='Last',
                    adjust=ADJUST_PREV,
                    end_time=end_date,
                    df=True
                )

                if price_data is None or len(price_data) < 15:
                    continue

                prices = price_data['close'].values
                volumes = price_data['volume'].values
                high = price_data['high'].values
                low = price_data['low'].values

                try:
                    fin_data = stk_get_financial_report(
                        symbol=symbol,
                        fields='pe_ttm,pb,ps,roe,eps,dividend_yield,net_profit',
                        report_type=REPORT_TYPE_QUARTER,
                        end_time=end_date,
                        df=True
                    )

                    pe_ttm = float(fin_data['pe_ttm'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['pe_ttm'].iloc[0]) else 40
                    pb = float(fin_data['pb'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['pb'].iloc[0]) else 4
                    ps = float(fin_data['ps'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['ps'].iloc[0]) else 3
                    roe = float(fin_data['roe'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['roe'].iloc[0]) else 0.08
                    eps = float(fin_data['eps'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['eps'].iloc[0]) else 0.5
                    dividend_yield = float(fin_data['dividend_yield'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['dividend_yield'].iloc[0]) else 0.02
                except Exception as e:
                    pe_ttm, pb, ps, roe, eps, dividend_yield = 40, 4, 3, 0.08, 0.5, 0.02

                try:
                    mv_data = stk_get_daily_mktvalue_pt(
                        symbols=[symbol],
                        fields='tot_mv,float_mv',
                        trade_date=end_date,
                        df=True
                    )
                    tot_mv = float(mv_data['tot_mv'].iloc[0]) if not mv_data.empty else 0
                    float_mv = float(mv_data['float_mv'].iloc[0]) if not mv_data.empty else 0
                except Exception as e:
                    tot_mv, float_mv = 0, 0

                short_days = self.config['strategy']['lookback_days_short']
                mid_days = self.config['strategy']['lookback_days_mid']
                long_days = self.config['strategy']['lookback_days_long']

                momentum_short = (prices[-1] / prices[-short_days-1]) - 1 if len(prices) > short_days else 0
                momentum_mid = (prices[-1] / prices[-mid_days-1]) - 1 if len(prices) > mid_days else 0
                momentum_long = (prices[-1] / prices[-long_days-1]) - 1 if len(prices) > long_days else 0

                reversal_5d = (prices[-6] / prices[-1]) - 1 if len(prices) > 6 else 0

                returns = np.log(prices[1:] / prices[:-1])
                volatility = returns.std() * np.sqrt(252)

                atr = self._calculate_atr(high, low, prices)
                avg_atr = np.mean(atr[-14:]) / prices[-1] if len(atr) >= 14 else 0.025

                volume_avg_5 = volumes[-5:].mean()
                volume_avg_20 = volumes[-20:-5].mean() if len(volumes) > 20 else volumes[-5:].mean()
                volume_ratio = volume_avg_5 / (volume_avg_20 + 1e-10)
                volume_trend = (volume_avg_5 / volume_avg_20) - 1

                rsi = self._calculate_rsi(prices)[-1] if len(prices) >= 14 else 50

                max_dd = self._calculate_max_drawdown(prices)

                beta = self._calculate_beta(prices, index_prices) if index_prices is not None else 1.0

                size_factor = -np.log(tot_mv) if tot_mv > 0 else 0

                result.append({
                    'symbol': symbol,
                    'name': symbol.split('.')[-1] if '.' in symbol else symbol,
                    'close': prices[-1],
                    'tot_mv': tot_mv,
                    'float_mv': float_mv,
                    'pe_ttm': pe_ttm,
                    'pb': pb,
                    'ps': ps,
                    'roe': roe,
                    'eps': eps,
                    'dividend_yield': dividend_yield,
                    'momentum_short': momentum_short,
                    'momentum_mid': momentum_mid,
                    'momentum_long': momentum_long,
                    'reversal_5d': reversal_5d,
                    'volatility': volatility,
                    'atr': avg_atr,
                    'volume_ratio': volume_ratio,
                    'volume_trend': volume_trend,
                    'rsi': rsi,
                    'max_drawdown': max_dd,
                    'beta': beta,
                    'size_factor': size_factor
                })

                if (idx + 1) % 50 == 0:
                    logger.info(f'已处理 {idx + 1}/{total} 只股票')

            except Exception as e:
                if idx % 20 == 0:
                    logger.warning(f'处理股票 {symbol} 失败: {e}')

        return pd.DataFrame(result)

    def factor_scoring(self, stock_df):
        """因子评分"""
        weights = self.adjust_weights_by_market(self.config['factors'], self.market_environment)

        stock_df['momentum_short_score'] = self._normalize(stock_df['momentum_short'])
        stock_df['momentum_mid_score'] = self._normalize(stock_df['momentum_mid'])
        stock_df['momentum_long_score'] = self._normalize(stock_df['momentum_long'])

        stock_df['reversal_score'] = self._normalize(stock_df['reversal_5d'])

        stock_df['value_pe_score'] = 1 - self._normalize(stock_df['pe_ttm'].clip(upper=120))
        stock_df['value_pb_score'] = 1 - self._normalize(stock_df['pb'].clip(upper=12))
        stock_df['value_ps_score'] = 1 - self._normalize(stock_df['ps'].clip(upper=25))
        stock_df['dividend_yield_score'] = self._normalize(stock_df['dividend_yield'])

        stock_df['volatility_score'] = 1 - self._normalize(stock_df['atr'])
        stock_df['quality_score'] = self._normalize(stock_df['roe'].clip(lower=-0.2, upper=0.5))
        stock_df['growth_score'] = self._normalize(stock_df['eps'].clip(lower=-1, upper=5))
        stock_df['beta_score'] = 1 - self._normalize(np.abs(stock_df['beta'] - 1.0))
        stock_df['size_score'] = self._normalize(stock_df['size_factor'])
        stock_df['volume_ratio_score'] = self._normalize(stock_df['volume_ratio'])

        stock_df['综合得分'] = (
            stock_df['momentum_short_score'] * weights.get('momentum_short_weight', 0.12) +
            stock_df['momentum_mid_score'] * weights.get('momentum_mid_weight', 0.15) +
            stock_df['momentum_long_score'] * weights.get('momentum_long_weight', 0.08) +
            stock_df['reversal_score'] * weights.get('reversal_weight', 0.05) +
            stock_df['value_pe_score'] * weights.get('valuation_pe_weight', 0.08) +
            stock_df['value_pb_score'] * weights.get('valuation_pb_weight', 0.06) +
            stock_df['value_ps_score'] * weights.get('valuation_ps_weight', 0.04) +
            stock_df['dividend_yield_score'] * weights.get('dividend_yield_weight', 0.06) +
            stock_df['volatility_score'] * weights.get('volatility_weight', 0.08) +
            stock_df['quality_score'] * weights.get('quality_roe_weight', 0.10) +
            stock_df['growth_score'] * weights.get('growth_weight', 0.08) +
            stock_df['beta_score'] * weights.get('beta_weight', 0.04) +
            stock_df['size_score'] * weights.get('size_weight', 0.02) +
            stock_df['volume_ratio_score'] * weights.get('volume_ratio_weight', 0.04)
        )

        return stock_df

    def filter_stocks(self, stock_df):
        """基本面筛选"""
        min_mv = self.config['strategy'].get('min_market_cap', 50)
        max_pe = self.config['strategy'].get('max_pe', 120)
        max_pb = self.config['strategy'].get('max_pb', 12)

        filtered = stock_df[
            (stock_df['tot_mv'] >= min_mv) &
            (stock_df['pe_ttm'] > 0) &
            (stock_df['pe_ttm'] <= max_pe) &
            (stock_df['pb'] <= max_pb) &
            (stock_df['roe'] > -0.2)
        ].copy()

        logger.info(f'基本面筛选后剩余 {len(filtered)} 只股票')
        return filtered

    def select_stocks(self, end_date=None):
        """执行选股"""
        if end_date is None:
            end_date = datetime.datetime.now().strftime('%Y-%m-%d')

        logger.info('=' * 60)
        logger.info('开始选股流程 V2.1')
        logger.info('=' * 60)

        best_sector = self.select_best_sector(end_date)

        try:
            constituents = stk_get_index_constituents(
                index=best_sector,
                trade_date=end_date
            )

            if constituents is None or len(constituents) == 0:
                logger.warning('无法获取成分股，尝试沪深300')
                best_sector = 'SHSE.000300'
                constituents = stk_get_index_constituents(
                    index=best_sector,
                    trade_date=end_date
                )

            symbols = list(constituents['symbol'])
            logger.info(f'获取 {self.config["indices"].get(best_sector, best_sector)} 成分股: {len(symbols)} 只')

        except Exception as e:
            logger.error(f'获取成分股失败: {e}')
            return pd.DataFrame()

        stock_df = self.get_stock_data(symbols, end_date)

        if stock_df.empty:
            logger.error('未获取到任何股票数据')
            return pd.DataFrame()

        filtered_df = self.filter_stocks(stock_df)
        scored_df = self.factor_scoring(filtered_df)

        if self.config['strategy'].get('enable_industry_neutral', True):
            try:
                stock_info = get_symbols(sec_type1=1010, symbols=scored_df['symbol'].tolist(), trade_date=end_date)
                if stock_info:
                    sector_map = {item['symbol']: item.get('sector_name', '未知') for item in stock_info}
                    scored_df['行业'] = scored_df['symbol'].map(sector_map)

                    final_selected = []
                    for sector in scored_df['行业'].unique():
                        sector_stocks = scored_df[scored_df['行业'] == sector]
                        take_num = min(len(sector_stocks), max(2, self.config['strategy']['holding_num'] // 8))
                        final_selected.extend(sector_stocks.nlargest(take_num, '综合得分')['symbol'].tolist())

                    if len(final_selected) < self.config['strategy']['holding_num']:
                        remaining = scored_df[~scored_df['symbol'].isin(final_selected)]
                        final_selected.extend(remaining.nlargest(
                            self.config['strategy']['holding_num'] - len(final_selected),
                            '综合得分'
                        )['symbol'].tolist())

                    selected = scored_df[scored_df['symbol'].isin(final_selected)]
                else:
                    selected = scored_df.nlargest(self.config['strategy']['holding_num'], '综合得分')
            except Exception as e:
                logger.warning(f'行业中性化处理失败，使用简单TopN: {e}')
                selected = scored_df.nlargest(self.config['strategy']['holding_num'], '综合得分')
        else:
            selected = scored_df.nlargest(self.config['strategy']['holding_num'], '综合得分')

        self.stock_pool = list(selected['symbol'])

        for _, row in selected.iterrows():
            self.stock_analysis[row['symbol']] = {
                'symbol': row['symbol'],
                'name': row['name'],
                'score': row['综合得分'],
                'close': row['close'],
                'momentum_short': row['momentum_short'],
                'momentum_mid': row['momentum_mid'],
                'momentum_long': row['momentum_long'],
                'reversal_5d': row['reversal_5d'],
                'pe_ttm': row['pe_ttm'],
                'pb': row['pb'],
                'ps': row['ps'],
                'roe': row['roe'],
                'eps': row['eps'],
                'dividend_yield': row['dividend_yield'],
                'volatility': row['volatility'],
                'atr': row['atr'],
                'beta': row['beta'],
                'size_factor': row['size_factor'],
                'volume_ratio': row['volume_ratio'],
                'volume_trend': row['volume_trend'],
                'rsi': row['rsi'],
                'max_drawdown': row['max_drawdown'],
                'tot_mv': row['tot_mv'],
                'float_mv': row['float_mv'],
                'market_environment': self.market_environment
            }

        logger.info('=' * 60)
        logger.info(f'选股完成，股票池共 {len(selected)} 只')
        logger.info(f'当前市场环境: {self.market_environment}')
        logger.info('=' * 60)

        return selected

    def generate_analysis_report(self, symbol):
        """生成个股分析报告"""
        if symbol not in self.stock_analysis:
            return None

        data = self.stock_analysis[symbol]
        report = []
        report.append('=' * 60)
        report.append(f'{data["name"]} ({data["symbol"]}) 个股分析报告 V2.1')
        report.append('=' * 60)
        report.append('')

        report.append('【一、基本信息】')
        report.append(f'  股票代码: {data["symbol"]}')
        report.append(f'  最新收盘价: {data["close"]:.2f} 元')
        report.append(f'  总市值: {data["tot_mv"]:.0f} 亿元')
        report.append(f'  流通市值: {data["float_mv"]:.0f} 亿元')
        report.append('')

        report.append('【二、估值指标】')
        report.append(f'  市盈率(PE_TTM): {data["pe_ttm"]:.2f}')
        report.append(f'  市净率(PB): {data["pb"]:.2f}')
        report.append(f'  市销率(PS): {data["ps"]:.2f}')
        report.append(f'  股息率: {data["dividend_yield"]*100:.2f}%')
        pe_status = '偏低' if data['pe_ttm'] < 20 else ('合理' if data['pe_ttm'] < 40 else '偏高')
        report.append(f'  估值评价: {pe_status}')
        report.append('')

        report.append('【三、质量指标】')
        report.append(f'  净资产收益率(ROE): {data["roe"]*100:.2f}%')
        report.append(f'  每股收益(EPS): {data["eps"]:.4f} 元')
        roe_status = '优秀' if data['roe'] > 0.15 else ('良好' if data['roe'] > 0.08 else '一般')
        report.append(f'  质量评价: {roe_status}')
        report.append('')

        report.append('【四、动量指标】')
        report.append(f'  短期动量(10日): {data["momentum_short"]*100:+.2f}%')
        report.append(f'  中期动量(20日): {data["momentum_mid"]*100:+.2f}%')
        report.append(f'  长期动量(60日): {data["momentum_long"]*100:+.2f}%')
        report.append(f'  短期反转(5日): {data["reversal_5d"]*100:+.2f}%')

        if data['momentum_mid'] > 0.12:
            momentum_status = '强势上涨'
        elif data['momentum_mid'] > 0:
            momentum_status = '温和上涨'
        elif data['momentum_mid'] > -0.1:
            momentum_status = '小幅回调'
        else:
            momentum_status = '明显下跌'
        report.append(f'  动量评价: {momentum_status}')
        report.append('')

        report.append('【五、风险指标】')
        report.append(f'  年化波动率: {data["volatility"]*100:.2f}%')
        report.append(f'  ATR指标: {data["atr"]*100:.2f}%')
        report.append(f'  贝塔系数(Beta): {data["beta"]:.2f}')
        report.append(f'  历史最大回撤: {data["max_drawdown"]*100:.2f}%')

        beta_status = '低贝塔' if data['beta'] < 0.8 else ('中贝塔' if data['beta'] < 1.2 else '高贝塔')
        report.append(f'  风险评价: {beta_status}')
        report.append('')

        report.append('【六、技术指标】')
        report.append(f'  RSI(14日): {data["rsi"]:.2f}')
        report.append(f'  量比: {data["volume_ratio"]:.2f}')
        report.append(f'  成交量趋势: {data["volume_trend"]*100:+.2f}%')

        if data['rsi'] > 75:
            rsi_status = '超买区域'
        elif data['rsi'] < 25:
            rsi_status = '超卖区域'
        else:
            rsi_status = '正常区间'
        report.append(f'  RSI评价: {rsi_status}')
        report.append('')

        report.append('【七、市场环境】')
        report.append(f'  当前环境: {data["market_environment"]}')
        env_descr = {'bull': '牛市（进取配置）', 'bear': '熊市（防御配置）', 
                    'volatile': '高波动（稳健配置）', 'neutral': '震荡市（均衡配置）'}
        report.append(f'  策略调整: {env_descr.get(data["market_environment"], "")}')
        report.append('')

        report.append('【八、综合评分】')
        report.append(f'  综合得分: {data["score"]:.4f}')
        if data['score'] > 0.72:
            overall = '强烈推荐'
        elif data['score'] > 0.55:
            overall = '推荐'
        elif data['score'] > 0.4:
            overall = '谨慎推荐'
        else:
            overall = '观望'
        report.append(f'  综合评价: {overall}')
        report.append('')

        report.append('=' * 60)
        return '\n'.join(report)

    def export_to_excel(self):
        """导出Excel报告"""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()

            ws_pool = wb.active
            ws_pool.title = '股票池汇总'

            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            ws_pool.append(['股票池汇总 V2.1'])
            ws_pool['A1'].font = Font(bold=True, size=14)
            ws_pool.merge_cells('A1:Y1')
            ws_pool.append([])
            ws_pool.append(['选股日期', datetime.datetime.now().strftime('%Y-%m-%d')])
            ws_pool.append(['股票数量', len(self.stock_pool)])
            ws_pool.append(['策略版本', self.config['strategy']['name']])
            ws_pool.append(['市场环境', self.market_environment])
            ws_pool.append([])

            headers = ['排名', '股票代码', '股票名称', '收盘价', '综合得分', 
                      'PE', 'PB', 'PS', '股息率(%)', 'ROE(%)',
                      '短期动量(%)', '中期动量(%)', '长期动量(%)', '5日反转(%)',
                      '波动率(%)', 'ATR(%)', 'Beta', 'RSI',
                      '量比', '成交量趋势(%)', '总市值(亿)', '最大回撤(%)']

            ws_pool.append(headers)
            for cell in ws_pool[ws_pool.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            sorted_stocks = sorted(
                self.stock_analysis.values(),
                key=lambda x: x['score'],
                reverse=True
            )

            for rank, stock in enumerate(sorted_stocks, 1):
                row_data = [
                    rank,
                    stock['symbol'],
                    stock['name'],
                    round(stock['close'], 2),
                    round(stock['score'], 4),
                    round(stock['pe_ttm'], 2),
                    round(stock['pb'], 2),
                    round(stock['ps'], 2),
                    round(stock['dividend_yield'] * 100, 2),
                    round(stock['roe'] * 100, 2),
                    round(stock['momentum_short'] * 100, 2),
                    round(stock['momentum_mid'] * 100, 2),
                    round(stock['momentum_long'] * 100, 2),
                    round(stock['reversal_5d'] * 100, 2),
                    round(stock['volatility'] * 100, 2),
                    round(stock['atr'] * 100, 2),
                    round(stock['beta'], 2),
                    round(stock['rsi'], 2),
                    round(stock['volume_ratio'], 2),
                    round(stock['volume_trend'] * 100, 2),
                    round(stock['tot_mv'], 0),
                    round(stock['max_drawdown'] * 100, 2)
                ]
                ws_pool.append(row_data)

            for row in ws_pool.iter_rows(min_row=ws_pool.min_row + 7, max_row=ws_pool.max_row):
                for cell in row:
                    cell.border = thin_border
                    if cell.column > 1:
                        cell.alignment = Alignment(horizontal='right')
                    if cell.column == 5 and cell.value:
                        if cell.value > 0.65:
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        elif cell.value > 0.45:
                            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            col_widths = [6, 14, 10, 9, 9, 8, 8, 8, 10, 10, 12, 12, 12, 12, 10, 9, 8, 8, 9, 12, 11, 12]
            for i, width in enumerate(col_widths, 1):
                if i <= 22:
                    ws_pool.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            for idx, stock in enumerate(sorted_stocks, 2):
                ws_report = wb.create_sheet(title=f'{stock["name"][:8]}')

                report = self.generate_analysis_report(stock['symbol'])
                lines = report.split('\n')

                title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

                for line in lines:
                    if '=' in line and line.strip().startswith('='):
                        ws_report.append([line])
                        ws_report[ws_report.max_row][0].fill = title_fill
                        ws_report[ws_report.max_row][0].font = Font(bold=True, size=12, color='FFFFFF')
                        ws_report.merge_cells(f'A{ws_report.max_row}:E{ws_report.max_row}')
                    elif '【' in line and '】' in line:
                        ws_report.append([line])
                        ws_report[ws_report.max_row][0].font = Font(bold=True, size=11)
                        ws_report[ws_report.max_row][0].fill = section_fill
                    elif line.strip():
                        ws_report.append([line])

                ws_report.column_dimensions['A'].width = 70

            ws_config = wb.create_sheet(title='策略配置')
            ws_config.append(['配置项', '值'])
            ws_config['A1'].font = Font(bold=True)
            ws_config['B1'].font = Font(bold=True)

            self._flatten_config(self.config, '', ws_config)
            ws_config.column_dimensions['A'].width = 35
            ws_config.column_dimensions['B'].width = 55

            filename = f'股票池V2_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(filename)
            logger.info(f'选股结果已导出到: {filename}')
            return filename

        except ImportError:
            logger.error('需要安装 openpyxl 库: pip install openpyxl')
        except Exception as e:
            logger.error(f'导出Excel失败: {e}')

    def _flatten_config(self, config, prefix, ws):
        """展开配置字典到Excel"""
        for key, value in config.items():
            new_key = f'{prefix}.{key}' if prefix else key
            if isinstance(value, dict):
                self._flatten_config(value, new_key, ws)
            else:
                ws.append([new_key, value])

    def print_summary(self):
        """打印摘要"""
        logger.info('=' * 80)
        logger.info('选股结果汇总 V2.1')
        logger.info('=' * 80)

        sorted_stocks = sorted(
            self.stock_analysis.values(),
            key=lambda x: x['score'],
            reverse=True
        )

        for rank, stock in enumerate(sorted_stocks, 1):
            logger.info(
                f'{rank:2d}. {stock["symbol"]} | 得分:{stock["score"]:.4f} | '
                f'PE:{stock["pe_ttm"]:.1f} | ROE:{stock["roe"]*100:.1f}% | '
                f'动量:{stock["momentum_mid"]*100:+.1f}% | '
                f'Beta:{stock["beta"]:.2f} | '
                f'市值:{stock["tot_mv"]:.0f}亿'
            )

            if rank <= 5:
                report = self.generate_analysis_report(stock['symbol'])
                if report:
                    print(report)
                    print()

        logger.info('=' * 80)
        logger.info(f'当前市场环境: {self.market_environment}')
        logger.info(f'共选出 {len(self.stock_pool)} 只股票')
        logger.info('=' * 80)


# ============ 使用示例 ============

# 简单使用（无需配置文件）
if __name__ == '__main__':
    import sys

    logger.info('=' * 60)
    logger.info('股票池选股器启动 V2.1')
    logger.info('提示：无需配置文件也能运行，使用内置默认配置')
    logger.info('=' * 60)

    # 初始化选股器（不指定配置文件，使用内置默认配置）
    selector = StockSelectorV2()

    # 执行选股
    selected = selector.select_stocks()

    if not selected.empty:
        selector.print_summary()
        filename = selector.export_to_excel()
        if filename:
            logger.info(f'Excel报告已生成: {filename}')
    else:
        logger.warning('选股结果为空')