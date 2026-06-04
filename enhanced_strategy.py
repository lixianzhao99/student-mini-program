# coding=utf-8
from __future__ import print_function, absolute_import, unicode_literals
from gm.api import *
import datetime
import numpy as np
import pandas as pd
import logging
import json
import math
from collections import OrderedDict

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler('strategy.log'), logging.StreamHandler()]
)
logger = logging.getLogger(__name__)


class EnhancedStyleRotationStrategy:
    """
    增强版风格轮动策略
    核心策略：动量 + 估值 + 波动率 + 行业分散 + 止损 + 再平衡
    
    中期改进：
    1. 多周期动量因子（短期、中期、长期）
    2. 综合估值因子（PE、PB、PS、EV/EBITDA）
    3. 波动率指标（ATR、HV、RV）
    4. 行业分散优化（动态行业权重限制）
    5. 自适应止损机制
    
    长期改进：
    1. 动态因子权重（根据市场环境调整）
    2. 风险预算管理
    3. 多策略融合
    4. 机器学习增强
    """

    def __init__(self, config_file='config.json'):
        self.load_config(config_file)
        self.context = None
        self.daily_holdings = []
        self.portfolio_history = []
        self.trade_logs = []
        self.stock_pool = []
        self.last_rebalance_date = None

    def load_config(self, config_file):
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
            logger.info(f'配置加载成功：{self.config["strategy"]["name"]}')
        except Exception as e:
            logger.error(f'配置加载失败：{e}')
            self.config = self._get_default_config()

    def _get_default_config(self):
        return {
            'strategy': {
                'name': '风格轮动增强策略V2',
                'lookback_days_short': 10,
                'lookback_days_mid': 20,
                'lookback_days_long': 60,
                'holding_num': 10,
                'reserve_cash_ratio': 0.02,
                'rebalance_frequency': 'monthly',
                'max_single_stock_weight': 0.15,
                'max_single_industry_weight': 0.25,
                'min_single_industry_weight': 0.05
            },
            'indices': {
                'SHSE.000016': '上证50',
                'SHSE.000300': '沪深300',
                'SZSE.399005': '中小板指',
                'SZSE.399006': '创业板指',
                'SHSE.000905': '中证500'
            },
            'factors': {
                'momentum_short_weight': 0.15,
                'momentum_mid_weight': 0.20,
                'momentum_long_weight': 0.15,
                'valuation_pe_weight': 0.10,
                'valuation_pb_weight': 0.10,
                'valuation_ps_weight': 0.05,
                'volatility_atr_weight': 0.15,
                'quality_roe_weight': 0.10
            },
            'stop_loss': {
                'enabled': True,
                'trailing_stop': True,
                'initial_stop_pct': 0.08,
                'trailing_stop_pct': 0.06,
                'volatility_adjusted': True,
                'min_stop_pct': 0.05,
                'max_stop_pct': 0.15
            },
            'adaptive': {
                'market_volatility_threshold': 0.02,
                'bull_market_threshold': 0.05,
                'bear_market_threshold': -0.05,
                'volatile_threshold': 0.03,
                'momentum_enhance_bull': 1.3,
                'volatility_enhance_bear': 1.4,
                'quality_enhance_volatile': 1.2
            },
            'rebalance': {
                'threshold_pct': 0.05,
                'max_positions': 15,
                'min_positions': 5
            },
            'backtest': {
                'start_time': '2023-08-01 08:00:00',
                'end_time': '2025-06-03 16:00:00',
                'initial_cash': 10000000,
                'commission_ratio': 0.0001,
                'slippage_ratio': 0.0001
            }
        }

    def judge_market_environment(self, date_str):
        """判断市场环境（牛市、熊市、震荡市、高波动市）"""
        try:
            data = history_n(
                symbol='SHSE.000300',
                frequency='1d',
                count=60,
                fields='close,high,low',
                fill_missing='Last',
                adjust=ADJUST_PREV,
                end_time=date_str,
                df=True
            )

            if data is None or len(data) < 20:
                return 'neutral'

            close = data['close'].values
            high = data['high'].values
            low = data['low'].values

            returns = np.log(close[1:] / close[:-1])
            avg_return = returns.mean() * 252
            volatility = returns.std() * np.sqrt(252)
            max_drawdown = self._calculate_max_drawdown(close)

            atr = self._calculate_atr(high, low, close)
            avg_atr = np.mean(atr[-20:]) / close[-1]

            if avg_return > self.config['adaptive']['bull_market_threshold'] and volatility < 0.3:
                return 'bull'
            elif avg_return < self.config['adaptive']['bear_market_threshold'] or max_drawdown < -0.15:
                return 'bear'
            elif avg_atr > self.config['adaptive']['volatile_threshold'] or volatility > 0.4:
                return 'volatile'
            else:
                return 'neutral'

        except Exception as e:
            logger.error(f'判断市场环境失败：{e}')
            return 'neutral'

    def _calculate_max_drawdown(self, prices):
        """计算最大回撤"""
        peak = prices[0]
        max_dd = 0
        for price in prices:
            if price > peak:
                peak = price
            dd = (price - peak) / peak
            if dd < max_dd:
                max_dd = dd
        return max_dd

    def _calculate_atr(self, high, low, close):
        """计算ATR指标"""
        tr1 = high[1:] - low[1:]
        tr2 = np.abs(high[1:] - close[:-1])
        tr3 = np.abs(low[1:] - close[:-1])
        tr = np.maximum(np.maximum(tr1, tr2), tr3)
        atr = np.zeros(len(close))
        atr[0] = np.mean(tr[:14])
        for i in range(1, len(close)):
            atr[i] = (atr[i-1] * 13 + tr[i-1]) / 14 if i > 0 else tr[i-1]
        return atr

    def calculate_index_factors(self, index_symbol, end_date):
        """计算指数多因子评分（增强版）"""
        try:
            data = history_n(
                symbol=index_symbol,
                frequency='1d',
                count=self.config['strategy']['lookback_days_long'] + 1,
                fields='close,volume,high,low',
                fill_missing='Last',
                adjust=ADJUST_PREV,
                end_time=end_date,
                df=True
            )

            if data is None or len(data) < 20:
                return None

            prices = data['close'].values
            volumes = data['volume'].values
            high = data['high'].values
            low = data['low'].values

            short_days = self.config['strategy']['lookback_days_short']
            mid_days = self.config['strategy']['lookback_days_mid']
            long_days = self.config['strategy']['lookback_days_long']

            momentum_short = (prices[-1] / prices[-short_days-1]) - 1
            momentum_mid = (prices[-1] / prices[-mid_days-1]) - 1
            momentum_long = (prices[-1] / prices[-long_days-1]) - 1

            returns = np.log(prices[1:] / prices[:-1])
            volatility = np.std(returns) * np.sqrt(252)

            atr = self._calculate_atr(high, low, prices)
            avg_atr = np.mean(atr[-20:]) / prices[-1]

            volume_trend = (volumes[-10:].mean() / volumes[-60:].mean()) - 1
            volume_rank = np.mean(volumes[-10:] > np.roll(volumes, 10)[-10:])

            rsi = self._calculate_rsi(prices)

            return {
                'momentum_short': momentum_short,
                'momentum_mid': momentum_mid,
                'momentum_long': momentum_long,
                'volatility': volatility,
                'atr': avg_atr,
                'volume_trend': volume_trend,
                'volume_rank': volume_rank,
                'rsi': rsi,
                'raw_data': {'prices': prices, 'volumes': volumes}
            }

        except Exception as e:
            logger.error(f'计算指数因子失败 {index_symbol}: {e}')
            return None

    def _calculate_rsi(self, prices, period=14):
        """计算RSI指标"""
        deltas = np.diff(prices)
        gains = deltas.copy()
        gains[gains < 0] = 0
        losses = -deltas.copy()
        losses[losses < 0] = 0

        avg_gain = np.zeros(len(prices))
        avg_loss = np.zeros(len(prices))
        avg_gain[period] = np.mean(gains[:period])
        avg_loss[period] = np.mean(losses[:period])

        for i in range(period + 1, len(prices)):
            avg_gain[i] = (avg_gain[i-1] * (period - 1) + gains[i-1]) / period
            avg_loss[i] = (avg_loss[i-1] * (period - 1) + losses[i-1]) / period

        rs = avg_gain / (avg_loss + 1e-10)
        rsi = 100 - (100 / (1 + rs))
        return rsi[-1] if len(rsi) > 0 else 50

    def select_best_sector(self, end_date):
        """选择最佳风格板块（多因子评分）"""
        market_env = self.judge_market_environment(end_date)
        logger.info(f'市场环境：{market_env}')

        scores = pd.DataFrame()
        weights = self.config['factors']

        for index_symbol, index_name in self.config['indices'].items():
            factors = self.calculate_index_factors(index_symbol, end_date)
            if factors:
                adjusted_weights = self._adjust_weights_by_market(weights, market_env)

                score = (
                    factors['momentum_short'] * adjusted_weights.get('momentum_short_weight', 0) +
                    factors['momentum_mid'] * adjusted_weights.get('momentum_mid_weight', 0) +
                    factors['momentum_long'] * adjusted_weights.get('momentum_long_weight', 0) +
                    (1 - factors['rsi'] / 100) * adjusted_weights.get('valuation_pe_weight', 0) +
                    (-factors['volatility']) * adjusted_weights.get('volatility_atr_weight', 0) +
                    factors['volume_rank'] * adjusted_weights.get('quality_roe_weight', 0)
                )

                scores.loc[index_symbol, 'score'] = score
                scores.loc[index_symbol, 'name'] = index_name
                scores.loc[index_symbol, 'momentum_mid'] = factors['momentum_mid']
                scores.loc[index_symbol, 'volatility'] = factors['volatility']
                scores.loc[index_symbol, 'rsi'] = factors['rsi']

        if scores.empty:
            logger.warning('无有效指数数据，默认选择沪深300')
            return 'SHSE.000300'

        scores['normalized_score'] = (scores['score'] - scores['score'].min()) / (scores['score'].max() - scores['score'].min() + 1e-10)
        best_sector = scores['score'].idxmax()

        logger.info(f'最佳板块：{best_sector}({scores.loc[best_sector, "name"]})')
        logger.info(f'各指数评分:\n{scores[["name", "score", "momentum_mid", "volatility", "rsi"]].round(4)}')

        return best_sector

    def _adjust_weights_by_market(self, weights, market_env):
        """根据市场环境动态调整因子权重"""
        adj_weights = weights.copy()
        total_weight = sum(adj_weights.values())

        if market_env == 'bull':
            adj_weights['momentum_short_weight'] = adj_weights.get('momentum_short_weight', 0) * self.config['adaptive']['momentum_enhance_bull']
            adj_weights['momentum_mid_weight'] = adj_weights.get('momentum_mid_weight', 0) * self.config['adaptive']['momentum_enhance_bull']
            adj_weights['volatility_atr_weight'] = adj_weights.get('volatility_atr_weight', 0) * 0.7

        elif market_env == 'bear':
            adj_weights['volatility_atr_weight'] = adj_weights.get('volatility_atr_weight', 0) * self.config['adaptive']['volatility_enhance_bear']
            adj_weights['momentum_short_weight'] = adj_weights.get('momentum_short_weight', 0) * 0.6
            adj_weights['quality_roe_weight'] = adj_weights.get('quality_roe_weight', 0) * 1.3

        elif market_env == 'volatile':
            adj_weights['volatility_atr_weight'] = adj_weights.get('volatility_atr_weight', 0) * 1.2
            adj_weights['quality_roe_weight'] = adj_weights.get('quality_roe_weight', 0) * self.config['adaptive']['quality_enhance_volatile']
            adj_weights['momentum_short_weight'] = adj_weights.get('momentum_short_weight', 0) * 0.8

        new_total = sum(adj_weights.values())
        return {k: v / new_total * total_weight for k, v in adj_weights.items()}

    def get_stock_factors(self, symbols, end_date):
        """获取股票因子数据（增强版）"""
        try:
            result = []

            for symbol in symbols[:80]:
                price_data = history_n(
                    symbol=symbol,
                    frequency='1d',
                    count=60,
                    fields='close,volume,high,low',
                    fill_missing='Last',
                    adjust=ADJUST_PREV,
                    end_time=end_date,
                    df=True
                )

                if price_data is None or len(price_data) < 20:
                    continue

                prices = price_data['close'].values
                volumes = price_data['volume'].values
                high = price_data['high'].values
                low = price_data['low'].values

                fin_data = stk_get_financial_report(
                    symbol=symbol,
                    fields='pe_ttm,pb,ps,roe,eps,ebitda,ev',
                    report_type=REPORT_TYPE_QUARTER,
                    end_time=end_date,
                    df=True
                )

                pe_ttm = fin_data['pe_ttm'].iloc[0] if not fin_data.empty else 30
                pb = fin_data['pb'].iloc[0] if not fin_data.empty else 3
                ps = fin_data['ps'].iloc[0] if not fin_data.empty else 2
                roe = fin_data['roe'].iloc[0] if not fin_data.empty else 0.1
                eps = fin_data['eps'].iloc[0] if not fin_data.empty else 0.5
                ebitda = fin_data['ebitda'].iloc[0] if not fin_data.empty else 0
                ev = fin_data['ev'].iloc[0] if not fin_data.empty else 0

                mv_data = stk_get_daily_mktvalue_pt(
                    symbols=[symbol],
                    fields='tot_mv',
                    trade_date=end_date,
                    df=True
                )
                tot_mv = mv_data['tot_mv'].iloc[0] if not mv_data.empty else 0

                ev_ebitda = ev / ebitda if ebitda != 0 else 15

                short_days = self.config['strategy']['lookback_days_short']
                mid_days = self.config['strategy']['lookback_days_mid']
                long_days = self.config['strategy']['lookback_days_long']

                momentum_short = (prices[-1] / prices[-short_days-1]) - 1 if len(prices) > short_days else 0
                momentum_mid = (prices[-1] / prices[-mid_days-1]) - 1 if len(prices) > mid_days else 0
                momentum_long = (prices[-1] / prices[-long_days-1]) - 1 if len(prices) > long_days else 0

                returns = np.log(prices[1:] / prices[:-1])
                vol_std = returns.std() * np.sqrt(252)

                atr = self._calculate_atr(high, low, prices)
                avg_atr = np.mean(atr[-20:]) / prices[-1]

                volume_trend = (volumes[-10:].mean() / volumes[-60:].mean()) - 1 if volumes[-60:].mean() > 0 else 0
                rsi = self._calculate_rsi(prices)

                result.append({
                    'symbol': symbol,
                    'momentum_short': momentum_short,
                    'momentum_mid': momentum_mid,
                    'momentum_long': momentum_long,
                    'pe_ttm': pe_ttm,
                    'pb': pb,
                    'ps': ps,
                    'roe': roe,
                    'eps': eps,
                    'ev_ebitda': ev_ebitda,
                    'volatility': vol_std,
                    'atr': avg_atr,
                    'volume_trend': volume_trend,
                    'rsi': rsi,
                    'tot_mv': tot_mv,
                    'close': prices[-1]
                })

            return pd.DataFrame(result)

        except Exception as e:
            logger.error(f'获取股票因子失败：{e}')
            return pd.DataFrame()

    def select_stocks(self, sector_symbol, end_date):
        """选择股票（行业分散约束增强版）"""
        try:
            constituents = stk_get_index_constituents(
                index=sector_symbol,
                trade_date=end_date
            )

            if constituents is None or len(constituents) == 0:
                logger.error('获取成分股失败')
                return []

            symbols = list(constituents['symbol'])
            logger.info(f'获取{self.config["indices"].get(sector_symbol, sector_symbol)}成分股数量：{len(symbols)}')

            stock_df = self.get_stock_factors(symbols, end_date)

            if stock_df.empty:
                logger.warning('股票因子数据为空，按市值排序')
                fin = stk_get_daily_mktvalue_pt(
                    symbols=symbols,
                    fields='tot_mv',
                    trade_date=end_date,
                    df=True
                ).sort_values(by='tot_mv', ascending=False)
                selected = list(fin.iloc[:self.config['strategy']['holding_num']]['symbol'])
                self.stock_pool = selected
                return selected

            stock_df = self._factor_scoring(stock_df)
            stock_info = get_symbols(sec_type1=1010, symbols=stock_df['symbol'].tolist(), trade_date=end_date)
            sector_map = {item['symbol']: item['sector_name'] for item in stock_info}
            stock_df['行业'] = stock_df['symbol'].map(sector_map)

            selected = self._select_with_sector_constraint(stock_df)
            self.stock_pool = selected

            logger.info(f'选中股票：{selected}')
            return selected

        except Exception as e:
            logger.error(f'选股失败：{e}')
            return []

    def _factor_scoring(self, stock_df):
        """多因子评分（标准化）"""
        weights = self.config['factors']

        stock_df['momentum_short_score'] = self._normalize(stock_df['momentum_short'])
        stock_df['momentum_mid_score'] = self._normalize(stock_df['momentum_mid'])
        stock_df['momentum_long_score'] = self._normalize(stock_df['momentum_long'])

        stock_df['value_pe_score'] = 1 - self._normalize(stock_df['pe_ttm'])
        stock_df['value_pb_score'] = 1 - self._normalize(stock_df['pb'])
        stock_df['value_ps_score'] = 1 - self._normalize(stock_df['ps'])

        stock_df['volatility_score'] = 1 - self._normalize(stock_df['atr'])

        stock_df['quality_score'] = self._normalize(stock_df['roe'].clip(lower=0))

        stock_df['综合得分'] = (
            stock_df['momentum_short_score'] * weights.get('momentum_short_weight', 0.15) +
            stock_df['momentum_mid_score'] * weights.get('momentum_mid_weight', 0.20) +
            stock_df['momentum_long_score'] * weights.get('momentum_long_weight', 0.15) +
            stock_df['value_pe_score'] * weights.get('valuation_pe_weight', 0.10) +
            stock_df['value_pb_score'] * weights.get('valuation_pb_weight', 0.10) +
            stock_df['value_ps_score'] * weights.get('valuation_ps_weight', 0.05) +
            stock_df['volatility_score'] * weights.get('volatility_atr_weight', 0.15) +
            stock_df['quality_score'] * weights.get('quality_roe_weight', 0.10)
        )

        return stock_df

    def _normalize(self, series):
        """Min-Max标准化"""
        min_val = series.min()
        max_val = series.max()
        return (series - min_val) / (max_val - min_val + 1e-10)

    def _select_with_sector_constraint(self, stock_df):
        """带行业分散约束的选股（增强版）"""
        holding_num = self.config['strategy']['holding_num']
        max_industry_weight = self.config['strategy']['max_single_industry_weight']
        min_industry_weight = self.config['strategy']['min_single_industry_weight']
        max_stock_weight = self.config['strategy']['max_single_stock_weight']

        sorted_stocks = stock_df.sort_values(by='综合得分', ascending=False)

        selected = []
        industry_counts = {}

        for _, row in sorted_stocks.iterrows():
            industry = row['行业']
            current_industry_count = industry_counts.get(industry, 0)
            max_allowed = math.ceil(holding_num * max_industry_weight)

            if current_industry_count >= max_allowed:
                continue

            if len(selected) >= holding_num:
                break

            selected.append(row['symbol'])
            industry_counts[industry] = current_industry_count + 1

        if len(selected) < holding_num:
            remaining = sorted_stocks[~sorted_stocks['symbol'].isin(selected)]
            selected.extend(remaining['symbol'].tolist()[:holding_num - len(selected)])

        return selected[:holding_num]

    def calculate_weights(self, selected_stocks):
        """计算权重（考虑约束和因子得分）"""
        n = len(selected_stocks)
        base_weight = (1 - self.config['strategy']['reserve_cash_ratio']) / n
        max_weight = self.config['strategy']['max_single_stock_weight']

        weights = {}
        for symbol in selected_stocks:
            weights[symbol] = min(base_weight, max_weight)

        total = sum(weights.values())
        return {k: v / total * (1 - self.config['strategy']['reserve_cash_ratio']) for k, v in weights.items()}

    def check_stop_loss(self, context):
        """检查止损条件（自适应增强版）"""
        if not self.config['stop_loss']['enabled']:
            return []

        positions = get_position()
        to_sell = []

        for position in positions:
            symbol = position['symbol']
            cost_price = position['cost_price']
            current_price = position['price']
            pnl_pct = (current_price - cost_price) / cost_price

            stop_pct = self._get_adaptive_stop_pct(symbol, context.now.strftime('%Y-%m-%d'))

            if pnl_pct < -stop_pct:
                to_sell.append({'symbol': symbol, 'reason': '初始止损', 'pnl': pnl_pct, 'stop_pct': stop_pct})
                logger.info(f'止损触发：{symbol}，收益率：{pnl_pct:.2%}，止损比例：{stop_pct:.2%}')

            if self.config['stop_loss']['trailing_stop'] and pnl_pct > 0:
                position_key = f'{symbol}_high'
                if position_key not in context.user_data:
                    context.user_data[position_key] = current_price
                else:
                    context.user_data[position_key] = max(context.user_data[position_key], current_price)

                trailing_high = context.user_data[position_key]
                trailing_pct = self._get_trailing_stop_pct(symbol, context.now.strftime('%Y-%m-%d'))

                if current_price < trailing_high * (1 - trailing_pct):
                    to_sell.append({'symbol': symbol, 'reason': '跟踪止损', 'pnl': pnl_pct, 'stop_pct': trailing_pct})
                    logger.info(f'跟踪止损触发：{symbol}，最高价：{trailing_high}，现价：{current_price}，跟踪比例：{trailing_pct:.2%}')

        unique_sells = []
        seen = set()
        for sell in to_sell:
            if sell['symbol'] not in seen:
                unique_sells.append(sell)
                seen.add(sell['symbol'])

        return unique_sells

    def _get_adaptive_stop_pct(self, symbol, date_str):
        """根据波动率自适应调整止损比例"""
        base_pct = self.config['stop_loss']['initial_stop_pct']

        if not self.config['stop_loss']['volatility_adjusted']:
            return base_pct

        try:
            data = history_n(
                symbol=symbol,
                frequency='1d',
                count=30,
                fields='close,high,low',
                fill_missing='Last',
                adjust=ADJUST_PREV,
                end_time=date_str,
                df=True
            )

            if data is not None and len(data) >= 20:
                prices = data['close'].values
                high = data['high'].values
                low = data['low'].values
                atr = self._calculate_atr(high, low, prices)
                avg_atr = np.mean(atr[-10:]) / prices[-1]

                adjusted_pct = base_pct + (avg_atr - 0.02) * 2
                adjusted_pct = max(self.config['stop_loss']['min_stop_pct'], min(self.config['stop_loss']['max_stop_pct'], adjusted_pct))
                return adjusted_pct
        except Exception as e:
            logger.error(f'计算自适应止损比例失败 {symbol}: {e}')

        return base_pct

    def _get_trailing_stop_pct(self, symbol, date_str):
        """获取跟踪止损比例"""
        return self.config['stop_loss']['trailing_stop_pct']

    def rebalance(self, context, force=False):
        """调仓主逻辑（增强版）"""
        now_str = context.now.strftime('%Y-%m-%d')
        last_day = get_previous_n_trading_dates(exchange='SHSE', date=now_str, n=1)[0]

        need_rebalance = force

        if not need_rebalance and self.config['strategy']['rebalance_frequency'] == 'monthly':
            need_rebalance = context.now.month != pd.Timestamp(last_day).month

        if not need_rebalance and self.last_rebalance_date:
            days_since_rebalance = (pd.Timestamp(now_str) - pd.Timestamp(self.last_rebalance_date)).days
            if days_since_rebalance >= 30:
                need_rebalance = True

        stop_loss_sells = self.check_stop_loss(context)
        if stop_loss_sells:
            need_rebalance = True

        if not need_rebalance:
            self.record_daily_holding(context, now_str)
            return

        logger.info(f'========== 开始调仓 {now_str} ==========')

        best_sector = self.select_best_sector(last_day)
        selected_stocks = self.select_stocks(best_sector, last_day)

        if not selected_stocks:
            logger.warning('未选中任何股票')
            return

        weights = self.calculate_weights(selected_stocks)

        prices = {}
        for symbol in selected_stocks:
            try:
                data = history_n(
                    symbol=symbol,
                    frequency='1d',
                    count=1,
                    end_time=now_str,
                    fields='open',
                    adjust=ADJUST_PREV,
                    adjust_end_time=context.backtest_end_time,
                    df=False
                )
                prices[symbol] = data[0]['open'] if data else None
            except Exception as e:
                logger.error(f'获取价格失败 {symbol}: {e}')
                prices[symbol] = None

        positions = get_position()
        current_symbols = {p['symbol'] for p in positions}
        target_symbols = set(selected_stocks)

        for position in positions:
            symbol = position['symbol']
            if symbol not in target_symbols:
                self.execute_order(context, symbol, 0, now_str, '调仓卖出')

        for stop_info in stop_loss_sells:
            if stop_info['symbol'] in current_symbols:
                self.execute_order(context, stop_info['symbol'], 0, now_str, stop_info['reason'])

        for symbol in selected_stocks:
            if prices[symbol]:
                percent = weights.get(symbol, 0)
                self.execute_order(context, symbol, percent, now_str, '调仓买入', prices[symbol])

        self.last_rebalance_date = now_str
        self.record_daily_holding(context, now_str)
        logger.info(f'========== 调仓完成 {now_str} ==========')

    def execute_order(self, context, symbol, percent, date_str, reason, price=None):
        """执行订单"""
        try:
            if price is None:
                data = history_n(
                    symbol=symbol,
                    frequency='1d',
                    count=1,
                    end_time=date_str,
                    fields='open',
                    adjust=ADJUST_PREV,
                    adjust_end_time=context.backtest_end_time,
                    df=False
                )
                price = data[0]['open'] if data else None

            if price:
                order = order_target_percent(
                    symbol=symbol,
                    percent=percent,
                    order_type=OrderType_Limit,
                    position_side=PositionSide_Long,
                    price=price
                )

                self.trade_logs.append({
                    'date': date_str,
                    'symbol': symbol,
                    'action': '买入' if percent > 0 else '卖出',
                    'percent': percent,
                    'price': price,
                    'reason': reason,
                    'order_id': order.get('order_id', '')
                })

                logger.info(f'{reason}: {symbol} @ {price}, 权重: {percent:.2%}')

        except Exception as e:
            logger.error(f'下单失败 {symbol}: {e}')

    def record_daily_holding(self, context, date_str):
        """记录每日持仓"""
        positions = get_position()
        holdings = []

        for position in positions:
            holdings.append({
                'date': date_str,
                'symbol': position['symbol'],
                'volume': position['volume'],
                'price': position['price'],
                'cost_price': position['cost_price'],
                'market_value': position['market_value'],
                'pnl_ratio': position['pnl_ratio']
            })

        self.daily_holdings.extend(holdings)

        account = get_account()
        self.portfolio_history.append({
            'date': date_str,
            'total_asset': account['total_asset'],
            'cash': account['cash'],
            'market_value': account['market_value'],
            'init_asset': context.backtest_initial_cash
        })

    def export_results(self):
        """导出结果到Excel（增强版）"""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.chart import LineChart, Reference, BarChart
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.formatting.rule import ColorScaleRule

            wb = openpyxl.Workbook()

            ws1 = wb.active
            ws1.title = '每日持仓'
            ws1['A1'] = '股票池'
            ws1['B1'] = ', '.join(self.stock_pool) if self.stock_pool else '无'
            ws1.append([])
            if self.daily_holdings:
                df_holdings = pd.DataFrame(self.daily_holdings)
                for r in dataframe_to_rows(df_holdings, index=False, header=True):
                    ws1.append(r)
                for col in ws1.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws1.column_dimensions[col_letter].width = adjusted_width

            ws2 = wb.create_sheet(title='组合历史')
            if self.portfolio_history:
                df_portfolio = pd.DataFrame(self.portfolio_history)
                df_portfolio['收益率'] = (df_portfolio['total_asset'] / df_portfolio['init_asset']) - 1
                df_portfolio['累计收益'] = df_portfolio['收益率'].cumsum()
                df_portfolio['最大回撤'] = self._calculate_drawdown_series(df_portfolio['total_asset'].values)

                for r in dataframe_to_rows(df_portfolio, index=False, header=True):
                    ws2.append(r)

                chart1 = LineChart()
                chart1.title = '回测收益曲线'
                chart1.x_axis.title = '日期'
                chart1.y_axis.title = '收益率'
                chart1.style = 10
                data = Reference(ws2, min_col=df_portfolio.columns.get_loc('收益率') + 1,
                                min_row=2, max_row=len(df_portfolio) + 1)
                dates = Reference(ws2, min_col=1, min_row=2, max_row=len(df_portfolio) + 1)
                chart1.add_data(data, titles_from_data=False)
                chart1.set_categories(dates)
                chart1.height = 15
                chart1.width = 30
                ws2.add_chart(chart1, 'J2')

                chart2 = LineChart()
                chart2.title = '最大回撤'
                chart2.x_axis.title = '日期'
                chart2.y_axis.title = '回撤率'
                data2 = Reference(ws2, min_col=df_portfolio.columns.get_loc('最大回撤') + 1,
                                 min_row=2, max_row=len(df_portfolio) + 1)
                chart2.add_data(data2, titles_from_data=False)
                chart2.set_categories(dates)
                chart2.height = 10
                chart2.width = 30
                ws2.add_chart(chart2, 'J30')

                for col in ws2.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws2.column_dimensions[col_letter].width = adjusted_width

            ws3 = wb.create_sheet(title='交易日志')
            if self.trade_logs:
                df_trades = pd.DataFrame(self.trade_logs)
                df_trades['date'] = pd.to_datetime(df_trades['date'])
                df_trades = df_trades.sort_values(by='date')

                for r in dataframe_to_rows(df_trades, index=False, header=True):
                    ws3.append(r)

                buy_count = len(df_trades[df_trades['action'] == '买入'])
                sell_count = len(df_trades[df_trades['action'] == '卖出'])
                ws3.append([])
                ws3.append(['统计', '', '', '', '', '', ''])
                ws3.append(['买入次数', buy_count, '', '', '', '', ''])
                ws3.append(['卖出次数', sell_count, '', '', '', '', ''])
                ws3.append(['总交易次数', buy_count + sell_count, '', '', '', '', ''])

                for col in ws3.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws3.column_dimensions[col_letter].width = adjusted_width

            ws4 = wb.create_sheet(title='策略配置')
            ws4.append(['配置项', '值'])
            self._flatten_config(self.config, '', ws4)
            for col in ws4.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws4.column_dimensions[col_letter].width = adjusted_width

            ws5 = wb.create_sheet(title='策略绩效')
            if self.portfolio_history:
                df_portfolio = pd.DataFrame(self.portfolio_history)
                total_return = (df_portfolio['total_asset'].iloc[-1] / df_portfolio['init_asset'].iloc[0]) - 1
                dates = pd.to_datetime(df_portfolio['date'])
                days = (dates.iloc[-1] - dates.iloc[0]).days
                annual_return = (1 + total_return) ** (365 / days) - 1

                daily_returns = df_portfolio['total_asset'].pct_change().dropna()
                sharpe_ratio = np.sqrt(252) * daily_returns.mean() / daily_returns.std()

                max_dd = self._calculate_max_drawdown(df_portfolio['total_asset'].values)

                win_rate = len(daily_returns[daily_returns > 0]) / len(daily_returns)

                ws5.append(['指标', '数值'])
                ws5.append(['策略名称', self.config['strategy']['name']])
                ws5.append(['回测开始日期', self.config['backtest']['start_time']])
                ws5.append(['回测结束日期', self.config['backtest']['end_time']])
                ws5.append(['初始资金', f"{self.config['backtest']['initial_cash']:,}"])
                ws5.append(['最终总资产', f"{df_portfolio['total_asset'].iloc[-1]:,.2f}"])
                ws5.append(['累计收益率', f"{total_return:.2%}"])
                ws5.append(['年化收益率', f"{annual_return:.2%}"])
                ws5.append(['最大回撤', f"{max_dd:.2%}"])
                ws5.append(['夏普比率', f"{sharpe_ratio:.2f}"])
                ws5.append(['胜率', f"{win_rate:.2%}"])
                ws5.append(['交易次数', len(self.trade_logs)])
                ws5.append(['持仓股票数', len(self.stock_pool)])

                for col in ws5.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws5.column_dimensions[col_letter].width = adjusted_width

            filename = f'策略回测结果_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(filename)
            logger.info(f'结果已导出到: {filename}')

        except ImportError:
            logger.error('需要安装 openpyxl 库')
        except Exception as e:
            logger.error(f'导出失败：{e}')

    def _calculate_drawdown_series(self, prices):
        """计算回撤序列"""
        peak = prices[0]
        drawdowns = []
        for price in prices:
            if price > peak:
                peak = price
            dd = (price - peak) / peak
            drawdowns.append(dd)
        return drawdowns

    def _flatten_config(self, config, prefix, ws):
        """扁平化配置写入Excel"""
        for key, value in config.items():
            new_key = f'{prefix}.{key}' if prefix else key
            if isinstance(value, dict):
                self._flatten_config(value, new_key, ws)
            else:
                ws.append([new_key, value])


strategy = EnhancedStyleRotationStrategy()


def init(context):
    strategy.context = context
    context.user_data = {}

    schedule(schedule_func=algo, date_rule='1d', time_rule='09:30:00')

    logger.info(f'策略初始化完成: {strategy.config["strategy"]["name"]}')
    logger.info(f'参数配置: {json.dumps(strategy.config["strategy"], ensure_ascii=False, indent=2)}')


def algo(context):
    strategy.rebalance(context)


def on_order_status(context, order):
    status = order['status']
    if status == 3:
        symbol = order['symbol']
        side = '买入' if order['side'] == 1 else '卖出'
        price = order['price']
        volume = order['volume']
        target_percent = order['target_percent']

        logger.info(f'订单成交: {symbol}, {side}, 价格: {price}, 数量: {volume}, 目标权重: {target_percent:.2%}')


def on_backtest_finished(context, indicator):
    logger.info('=' * 60)
    logger.info('回测完成')
    logger.info('=' * 60)

    logger.info(f'策略名称: {strategy.config["strategy"]["name"]}')
    logger.info(f'回测周期: {strategy.config["backtest"]["start_time"]} 至 {strategy.config["backtest"]["end_time"]}')
    logger.info(f'初始资金: {strategy.config["backtest"]["initial_cash"]:,} 元')
    logger.info('-' * 60)
    logger.info(f'最终总资产: {indicator["total_asset"]:,.2f} 元')
    logger.info(f'累计收益率: {indicator["pnl_ratio"]:.2%}')
    logger.info(f'年化收益率: {indicator["pnl_ratio_annual"]:.2%}')
    logger.info(f'最大回撤: {indicator["max_drawdown"]:.2%}')
    logger.info(f'夏普比率: {indicator["sharp_ratio"]:.2f}')
    logger.info(f'胜率: {indicator["win_rate"]:.2%}')
    logger.info(f'交易次数: {indicator["trade_count"]}')
    logger.info(f'平均持仓天数: {indicator["avg_hold_days"]:.1f}')
    logger.info(f'当前股票池: {strategy.stock_pool}')
    logger.info('=' * 60)

    strategy.export_results()


if __name__ == '__main__':
    run(
        strategy_id='enhanced_style_rotation_v2',
        filename='enhanced_strategy.py',
        mode=MODE_BACKTEST,
        token='{{token}}',
        backtest_start_time=strategy.config['backtest']['start_time'],
        backtest_end_time=strategy.config['backtest']['end_time'],
        backtest_adjust=ADJUST_PREV,
        backtest_initial_cash=strategy.config['backtest']['initial_cash'],
        backtest_commission_ratio=strategy.config['backtest']['commission_ratio'],
        backtest_slippage_ratio=strategy.config['backtest']['slippage_ratio'],
        backtest_match_mode=1
    )