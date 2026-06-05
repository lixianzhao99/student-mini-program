# coding=utf-8
"""
股票池分析器 V2 - Baostock版本
- 使用Baostock获取免费数据
- 支持用户自定义股票列表
- 本地Jupyter运行
- 核心逻辑：动量 + 估值 + 质量 + 波动率 + 动态权重

作者：量化学习项目
版本：V2.0 (Baostock版)
"""

import baostock as bs
import pandas as pd
import numpy as np
import datetime
import warnings
warnings.filterwarnings('ignore')

# 尝试导入openpyxl，如果失败则提示
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ 提示：需要安装 openpyxl 才能导出Excel")
    print("   运行命令：pip install openpyxl")


class StockAnalyzer:
    """
    股票分析器 V2.0
    核心功能：
    1. 多因子评分（动量、估值、质量、波动率）
    2. 动态权重调整（根据市场环境）
    3. 个股详细分析报告
    4. Excel导出
    """
    
    def __init__(self):
        """初始化分析器"""
        # 登录Baostock
        lg = bs.login()
        if lg.error_code != '0':
            raise Exception(f"Baostock登录失败: {lg.error_msg}")
        print(f"✅ Baostock登录成功")
        
        # 策略配置
        self.config = {
            'strategy': {
                'name': '多因子分析策略V2.0',
                'lookback_days_short': 10,
                'lookback_days_mid': 20,
                'lookback_days_long': 60,
            },
            'factors': {
                'momentum_short_weight': 0.12,
                'momentum_mid_weight': 0.15,
                'momentum_long_weight': 0.08,
                'valuation_pe_weight': 0.08,
                'valuation_pb_weight': 0.06,
                'quality_roe_weight': 0.10,
                'growth_weight': 0.08,
                'volatility_weight': 0.08,
                'beta_weight': 0.04,
                'reversal_weight': 0.05,
                'volume_weight': 0.04,
                'dividend_weight': 0.06,
                'size_weight': 0.02
            },
            'market_adjust': {
                'bull': {'momentum': 1.4, 'value': 0.9, 'quality': 0.9, 'volatility': 0.8},
                'bear': {'momentum': 0.7, 'value': 1.3, 'quality': 1.4, 'volatility': 1.4},
                'volatile': {'momentum': 0.85, 'value': 1.1, 'quality': 1.2, 'volatility': 1.3},
                'neutral': {'momentum': 1.0, 'value': 1.0, 'quality': 1.0, 'volatility': 1.0}
            }
        }
        
        self.stock_analysis = {}
        self.market_environment = 'neutral'
        self.stock_pool = []
    
    def logout(self):
        """登出Baostock"""
        bs.logout()
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.logout()
        return False
    
    def judge_market_environment(self, date_str):
        """
        判断市场环境
        根据沪深300指数判断当前市场状态
        """
        try:
            # 获取沪深300历史数据
            rs = bs.query_history_k_data_plus(
                "sh.000300",
                "date,close,volume",
                start_date=(pd.to_datetime(date_str) - datetime.timedelta(days=45)).strftime('%Y-%m-%d'),
                end_date=date_str,
                frequency="d"
            )
            
            data_list = []
            while (rs.error_code == '0') & rs.next():
                data_list.append(rs.get_row_data())
            
            if len(data_list) < 20:
                return 'neutral'
            
            df = pd.DataFrame(data_list, columns=['date', 'close', 'volume'])
            df['close'] = df['close'].astype(float)
            df['volume'] = df['volume'].astype(float)
            
            # 计算指标
            returns = df['close'].pct_change().dropna()
            avg_return = returns.mean()
            volatility = returns.std()
            volume_trend = df['volume'].iloc[-5:].mean() / df['volume'].iloc[-20:-5].mean() - 1
            
            # 判断环境
            if avg_return > 0.002 and volume_trend > 0.1:
                return 'bull'
            elif avg_return < -0.002:
                return 'bear'
            elif volatility > 0.025:
                return 'volatile'
            else:
                return 'neutral'
                
        except Exception as e:
            print(f"⚠️ 判断市场环境失败: {e}")
            return 'neutral'
    
    def _normalize(self, series):
        """Min-Max标准化"""
        min_val = series.min()
        max_val = series.max()
        if max_val == min_val:
            return pd.Series([0.5] * len(series), index=series.index)
        return (series - min_val) / (max_val - min_val + 1e-10)
    
    def adjust_weights_by_market(self, weights, market_env):
        """根据市场环境调整因子权重"""
        adj_weights = weights.copy()
        adjust_params = self.config['market_adjust'].get(market_env, {})
        
        for factor_type, multiplier in adjust_params.items():
            if factor_type == 'momentum':
                adj_weights['momentum_short_weight'] *= multiplier
                adj_weights['momentum_mid_weight'] *= multiplier
                adj_weights['momentum_long_weight'] *= multiplier
            elif factor_type == 'value':
                adj_weights['valuation_pe_weight'] *= multiplier
                adj_weights['valuation_pb_weight'] *= multiplier
            elif factor_type == 'quality':
                adj_weights['quality_roe_weight'] *= multiplier
                adj_weights['growth_weight'] *= multiplier
            elif factor_type == 'volatility':
                adj_weights['volatility_weight'] *= multiplier
        
        # 归一化
        total = sum(adj_weights.values())
        return {k: v / total for k, v in adj_weights.items()}
    
    def get_stock_data(self, symbol, end_date):
        """
        获取单只股票数据
        symbol: 股票代码，如 'sh.600519' 或 'sz.000001'
        """
        result = {}
        
        try:
            # 获取K线数据（包含价格和成交量）
            rs = bs.query_history_k_data_plus(
                symbol,
                "date,code,open,high,low,close,volume,amount,turn",
                start_date=(pd.to_datetime(end_date) - datetime.timedelta(days=90)).strftime('%Y-%m-%d'),
                end_date=end_date,
                frequency="d"
            )
            
            data_list = []
            while (rs.error_code == '0') & rs.next():
                data_list.append(rs.get_row_data())
            
            if len(data_list) < 30:
                print(f"⚠️ {symbol} 数据不足")
                return None
            
            df = pd.DataFrame(data_list, columns=['date', 'code', 'open', 'high', 'low', 'close', 'volume', 'amount', 'turn'])
            
            # 转换为数值类型
            for col in ['open', 'high', 'low', 'close', 'volume', 'amount', 'turn']:
                df[col] = pd.to_numeric(df[col], errors='coerce')
            
            df = df.dropna()
            
            if len(df) < 30:
                return None
            
            result['df'] = df
            result['name'] = symbol
            
            # ========== 计算各因子 ==========
            
            # 1. 动量因子
            short_days = self.config['strategy']['lookback_days_short']
            mid_days = self.config['strategy']['lookback_days_mid']
            long_days = self.config['strategy']['lookback_days_long']
            
            prices = df['close'].values
            
            result['momentum_short'] = (prices[-1] / prices[-short_days-1]) - 1 if len(prices) > short_days else 0
            result['momentum_mid'] = (prices[-1] / prices[-mid_days-1]) - 1 if len(prices) > mid_days else 0
            result['momentum_long'] = (prices[-1] / prices[-long_days-1]) - 1 if len(prices) > long_days else 0
            
            # 2. 波动率因子（ATR）
            high = df['high'].values
            low = df['low'].values
            close = df['close'].values
            
            tr1 = high[1:] - low[1:]
            tr2 = np.abs(high[1:] - close[:-1])
            tr3 = np.abs(low[1:] - close[:-1])
            tr = np.maximum(np.maximum(tr1, tr2), tr3)
            atr = np.zeros(len(close))
            if len(tr) >= 14:
                atr[14] = np.mean(tr[:14])
                for i in range(15, len(close)):
                    atr[i] = (atr[i-1] * 13 + tr[i-1]) / 14
            result['atr'] = np.mean(atr[-14:]) / prices[-1] if len(atr) >= 14 else 0.025
            
            # 年化波动率
            returns = np.diff(np.log(prices))
            result['volatility'] = np.std(returns) * np.sqrt(252)
            
            # 3. 反转因子（5日）
            result['reversal_5d'] = (prices[-6] / prices[-1]) - 1 if len(prices) > 6 else 0
            
            # 4. 量比
            vol = df['volume'].values
            vol_avg_5 = np.mean(vol[-5:])
            vol_avg_20 = np.mean(vol[-20:-5]) if len(vol) > 20 else vol_avg_5
            result['volume_ratio'] = vol_avg_5 / (vol_avg_20 + 1e-10)
            result['volume_trend'] = (vol_avg_5 / vol_avg_20) - 1
            
            # 5. RSI
            deltas = np.diff(prices)
            gains = deltas.copy()
            gains[gains < 0] = 0
            losses = -deltas.copy()
            losses[losses < 0] = 0
            avg_gain = np.zeros(len(prices))
            avg_loss = np.zeros(len(prices))
            if len(gains) >= 14:
                avg_gain[14] = np.mean(gains[:14])
                avg_loss[14] = np.mean(losses[:14])
                for i in range(15, len(prices)):
                    avg_gain[i] = (avg_gain[i-1] * 13 + gains[i-1]) / 14
                    avg_loss[i] = (avg_loss[i-1] * 13 + losses[i-1]) / 14
            rs_values = avg_gain / (avg_loss + 1e-10)
            result['rsi'] = 100 - (100 / (1 + rs_values))
            result['rsi'] = result['rsi'][-1] if len(result['rsi']) >= 14 else 50
            
            # 6. 最大回撤
            peak = prices[0]
            max_dd = 0
            for price in prices:
                if price > peak:
                    peak = price
                dd = (price - peak) / peak
                if dd < max_dd:
                    max_dd = dd
            result['max_drawdown'] = max_dd
            
            # 基本信息
            result['close'] = prices[-1]
            result['high_60d'] = np.max(prices[-60:]) if len(prices) >= 60 else np.max(prices)
            result['low_60d'] = np.min(prices[-60:]) if len(prices) >= 60 else np.min(prices)
            
            # 获取财务数据
            code = symbol.replace('sh.', '').replace('sz.', '')
            
            # 市值数据
            rs_mv = bs.query_daily_data(code, end_date, "year")
            mv_data = []
            while (rs_mv.error_code == '0') & rs_mv.next():
                mv_data.append(rs_mv.get_row_data())
            
            if mv_data:
                result['tot_mv'] = float(mv_data[0][3]) if len(mv_data[0]) > 3 else 0
            else:
                result['tot_mv'] = 0
            
            # 财务指标
            rs_fina = bs.query_profit_data(code=code, year=pd.to_datetime(end_date).year, quarter=4)
            fina_list = []
            while (rs_fina.error_code == '0') & rs_fina.next():
                fina_list.append(rs_fina.get_row_data())
            
            if fina_list:
                row = fina_list[0]
                result['pe'] = float(row[3]) if row[3] and row[3] != '' else 30
                result['pb'] = float(row[4]) if row[4] and row[4] != '' else 3
                result['roe'] = float(row[4]) / 100 if row[4] and row[4] != '' else 0.1
            else:
                result['pe'] = 30
                result['pb'] = 3
                result['roe'] = 0.1
            
            # 股息率（简化版）
            result['dividend_yield'] = 0.02  # 默认值
            
            # Beta系数（简化版，相对于上证指数）
            try:
                rs_idx = bs.query_history_k_data_plus(
                    "sh.000001",
                    "close",
                    start_date=(pd.to_datetime(end_date) - datetime.timedelta(days=90)).strftime('%Y-%m-%d'),
                    end_date=end_date,
                    frequency="d"
                )
                idx_list = []
                while (rs_idx.error_code == '0') & rs_idx.next():
                    idx_list.append(rs_idx.get_row_data())
                
                if len(idx_list) >= 60:
                    idx_prices = np.array([float(x[0]) for x in idx_list])
                    idx_returns = np.diff(np.log(idx_prices))
                    cov = np.cov(returns[-60:], idx_returns[-60:])[0, 1]
                    idx_var = np.var(idx_returns[-60:])
                    result['beta'] = np.clip(cov / (idx_var + 1e-10), 0.5, 1.5)
                else:
                    result['beta'] = 1.0
            except:
                result['beta'] = 1.0
            
            return result
            
        except Exception as e:
            print(f"⚠️ 获取 {symbol} 数据失败: {e}")
            return None
    
    def analyze_stocks(self, symbols, end_date=None):
        """
        分析多只股票
        symbols: 股票代码列表，如 ['sh.600519', 'sh.601318', 'sz.000858']
        end_date: 分析日期，默认今天
        """
        if end_date is None:
            end_date = datetime.datetime.now().strftime('%Y-%m-%d')
        
        print("\n" + "=" * 70)
        print(f"📊 股票分析器 V2.0")
        print(f"📅 分析日期: {end_date}")
        print(f"📈 股票数量: {len(symbols)}")
        print("=" * 70)
        
        # 判断市场环境
        self.market_environment = self.judge_market_environment(end_date)
        env_desc = {'bull': '牛市 🐂', 'bear': '熊市 🐻', 'volatile': '高波动 ⚡', 'neutral': '震荡 📊'}
        print(f"\n🌍 市场环境: {env_desc.get(self.market_environment, '未知')}")
        
        # 调整后的权重
        adjusted_weights = self.adjust_weights_by_market(self.config['factors'], self.market_environment)
        print("\n⚙️ 因子权重（已根据市场环境调整）:")
        print(f"   动量因子: {adjusted_weights.get('momentum_short_weight', 0) + adjusted_weights.get('momentum_mid_weight', 0) + adjusted_weights.get('momentum_long_weight', 0):.1%}")
        print(f"   估值因子: {adjusted_weights.get('valuation_pe_weight', 0) + adjusted_weights.get('valuation_pb_weight', 0):.1%}")
        print(f"   质量因子: {adjusted_weights.get('quality_roe_weight', 0) + adjusted_weights.get('growth_weight', 0):.1%}")
        print(f"   风险因子: {adjusted_weights.get('volatility_weight', 0) + adjusted_weights.get('beta_weight', 0):.1%}")
        
        # 获取所有股票数据
        print(f"\n📥 正在获取股票数据...")
        stock_data = []
        
        for i, symbol in enumerate(symbols, 1):
            print(f"   [{i}/{len(symbols)}] 获取 {symbol}...", end=' ')
            data = self.get_stock_data(symbol, end_date)
            if data:
                data['symbol'] = symbol
                stock_data.append(data)
                print("✅")
            else:
                print("❌")
        
        if not stock_data:
            print("\n❌ 未获取到任何股票数据")
            return None
        
        print(f"\n✅ 成功获取 {len(stock_data)} 只股票数据")
        
        # 转换为DataFrame进行评分
        df = pd.DataFrame(stock_data)
        
        # ========== 因子评分 ==========
        print("\n📈 正在计算因子评分...")
        
        # 动量评分
        df['mom_short_score'] = self._normalize(df['momentum_short'])
        df['mom_mid_score'] = self._normalize(df['momentum_mid'])
        df['mom_long_score'] = self._normalize(df['momentum_long'])
        
        # 估值评分（越低越好）
        df['pe_score'] = 1 - self._normalize(df['pe'].clip(upper=100))
        df['pb_score'] = 1 - self._normalize(df['pb'].clip(upper=10))
        
        # 质量评分
        df['roe_score'] = self._normalize(df['roe'].clip(lower=0, upper=0.5))
        
        # 波动率评分（越低越好）
        df['vol_score'] = 1 - self._normalize(df['atr'])
        
        # Beta评分（接近1.0越好）
        df['beta_score'] = 1 - self._normalize(np.abs(df['beta'] - 1.0))
        
        # 反转评分
        df['reversal_score'] = self._normalize(df['reversal_5d'])
        
        # 计算综合得分
        df['综合得分'] = (
            df['mom_short_score'] * adjusted_weights.get('momentum_short_weight', 0.12) +
            df['mom_mid_score'] * adjusted_weights.get('momentum_mid_weight', 0.15) +
            df['mom_long_score'] * adjusted_weights.get('momentum_long_weight', 0.08) +
            df['pe_score'] * adjusted_weights.get('valuation_pe_weight', 0.08) +
            df['pb_score'] * adjusted_weights.get('valuation_pb_weight', 0.06) +
            df['roe_score'] * adjusted_weights.get('quality_roe_weight', 0.10) +
            df['vol_score'] * adjusted_weights.get('volatility_weight', 0.08) +
            df['beta_score'] * adjusted_weights.get('beta_weight', 0.04) +
            df['reversal_score'] * adjusted_weights.get('reversal_weight', 0.05)
        )
        
        # 按得分排序
        df = df.sort_values('综合得分', ascending=False).reset_index(drop=True)
        
        # 保存结果
        self.stock_analysis = df.set_index('symbol').to_dict('index')
        self.stock_pool = list(df['symbol'])
        
        # 输出汇总
        self._print_summary()
        
        return df
    
    def _print_summary(self):
        """打印分析汇总"""
        print("\n" + "=" * 70)
        print("📊 股票分析汇总")
        print("=" * 70)
        
        print(f"\n{'排名':^4} {'代码':^12} {'收盘价':^10} {'综合得分':^10} {'PE':^8} {'ROE':^10} {'动量(20日)':^12} {'Beta':^8}")
        print("-" * 70)
        
        for i, (symbol, data) in enumerate(self.stock_analysis.items(), 1):
            print(f"{i:^4} {symbol:^12} {data['close']:^10.2f} {data['综合得分']:^10.4f} "
                  f"{data['pe']:^8.2f} {data['roe']*100:^10.2f}% {data['momentum_mid']*100:^12.2f}% {data['beta']:^8.2f}")
        
        print("-" * 70)
    
    def generate_report(self, symbol):
        """生成单只股票详细分析报告"""
        if symbol not in self.stock_analysis:
            return None
        
        data = self.stock_analysis[symbol]
        
        report = []
        report.append("\n" + "=" * 70)
        report.append(f"📈 {symbol} 个股详细分析报告")
        report.append("=" * 70)
        
        # 基本信息
        report.append("\n【一、基本信息】")
        report.append(f"   股票代码: {symbol}")
        report.append(f"   当前价格: {data['close']:.2f} 元")
        report.append(f"   60日最高: {data['high_60d']:.2f} 元")
        report.append(f"   60日最低: {data['low_60d']:.2f} 元")
        report.append(f"   总市值: {data['tot_mv']:.2f} 亿元" if data['tot_mv'] > 0 else "   总市值: 暂无数据")
        
        # 估值指标
        report.append("\n【二、估值指标】")
        report.append(f"   市盈率(PE): {data['pe']:.2f}")
        report.append(f"   市净率(PB): {data['pb']:.2f}")
        pe_status = '偏低' if data['pe'] < 20 else ('合理' if data['pe'] < 40 else '偏高')
        report.append(f"   估值评价: {pe_status}")
        
        # 质量指标
        report.append("\n【三、质量指标】")
        report.append(f"   净资产收益率(ROE): {data['roe']*100:.2f}%")
        roe_status = '优秀' if data['roe'] > 0.15 else ('良好' if data['roe'] > 0.08 else '一般')
        report.append(f"   质量评价: {roe_status}")
        
        # 动量指标
        report.append("\n【四、动量指标】")
        report.append(f"   短期动量(10日): {data['momentum_short']*100:+.2f}%")
        report.append(f"   中期动量(20日): {data['momentum_mid']*100:+.2f}%")
        report.append(f"   长期动量(60日): {data['momentum_long']*100:+.2f}%")
        report.append(f"   短期反转(5日): {data['reversal_5d']*100:+.2f}%")
        
        if data['momentum_mid'] > 0.12:
            mom_status = '强势上涨'
        elif data['momentum_mid'] > 0:
            mom_status = '温和上涨'
        elif data['momentum_mid'] > -0.1:
            mom_status = '小幅回调'
        else:
            mom_status = '明显下跌'
        report.append(f"   动量评价: {mom_status}")
        
        # 风险指标
        report.append("\n【五、风险指标】")
        report.append(f"   年化波动率: {data['volatility']*100:.2f}%")
        report.append(f"   ATR指标: {data['atr']*100:.2f}%")
        report.append(f"   Beta系数: {data['beta']:.2f}")
        report.append(f"   历史最大回撤: {data['max_drawdown']*100:.2f}%")
        
        beta_status = '低贝塔' if data['beta'] < 0.8 else ('中贝塔' if data['beta'] < 1.2 else '高贝塔')
        report.append(f"   风险评价: {beta_status}")
        
        # 技术指标
        report.append("\n【六、技术指标】")
        report.append(f"   RSI(14日): {data['rsi']:.2f}")
        report.append(f"   量比: {data['volume_ratio']:.2f}")
        report.append(f"   成交量趋势: {data['volume_trend']*100:+.2f}%")
        
        if data['rsi'] > 75:
            rsi_status = '超买区域'
        elif data['rsi'] < 25:
            rsi_status = '超卖区域'
        else:
            rsi_status = '正常区间'
        report.append(f"   RSI评价: {rsi_status}")
        
        # 综合评分
        report.append("\n【七、综合评分】")
        report.append(f"   综合得分: {data['综合得分']:.4f}")
        if data['综合得分'] > 0.7:
            overall = '强烈推荐 ⭐⭐⭐'
        elif data['综合得分'] > 0.55:
            overall = '推荐 ⭐⭐'
        elif data['综合得分'] > 0.4:
            overall = '谨慎推荐 ⭐'
        else:
            overall = '观望'
        report.append(f"   综合评价: {overall}")
        
        report.append("\n" + "=" * 70)
        
        return '\n'.join(report)
    
    def print_all_reports(self):
        """打印所有股票的分析报告"""
        for symbol in self.stock_pool:
            print(self.generate_report(symbol))
    
    def export_to_excel(self, filename=None):
        """导出分析结果到Excel"""
        if not EXCEL_AVAILABLE:
            print("\n❌ 无法导出Excel：缺少 openpyxl 库")
            print("   运行命令：pip install openpyxl")
            return None
        
        if not self.stock_analysis:
            print("\n❌ 没有分析数据可导出")
            return None
        
        if filename is None:
            filename = f"股票分析_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        try:
            wb = Workbook()
            
            # Sheet1: 股票池汇总
            ws_summary = wb.active
            ws_summary.title = '股票池汇总'
            
            # 标题
            ws_summary['A1'] = '股票池分析汇总'
            ws_summary['A1'].font = Font(bold=True, size=14)
            ws_summary.merge_cells('A1:N1')
            
            ws_summary.append([])
            ws_summary.append(['分析日期', datetime.datetime.now().strftime('%Y-%m-%d')])
            ws_summary.append(['市场环境', self.market_environment])
            ws_summary.append(['股票数量', len(self.stock_analysis)])
            ws_summary.append([])
            
            # 表头
            headers = ['排名', '股票代码', '收盘价', '综合得分', 'PE', 'PB', 'ROE(%)', 
                      '短期动量(%)', '中期动量(%)', '长期动量(%)', '波动率(%)', 'Beta', 'RSI', '最大回撤(%)']
            ws_summary.append(headers)
            
            # 设置表头样式
            for cell in ws_summary[ws_summary.max_row]:
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 数据
            for rank, symbol in enumerate(self.stock_pool, 1):
                data = self.stock_analysis[symbol]
                ws_summary.append([
                    rank,
                    symbol,
                    round(data['close'], 2),
                    round(data['综合得分'], 4),
                    round(data['pe'], 2),
                    round(data['pb'], 2),
                    round(data['roe'] * 100, 2),
                    round(data['momentum_short'] * 100, 2),
                    round(data['momentum_mid'] * 100, 2),
                    round(data['momentum_long'] * 100, 2),
                    round(data['volatility'] * 100, 2),
                    round(data['beta'], 2),
                    round(data['rsi'], 2),
                    round(data['max_drawdown'] * 100, 2)
                ])
            
            # 设置列宽
            col_widths = [6, 12, 10, 10, 8, 8, 10, 12, 12, 12, 12, 8, 8, 12]
            for i, width in enumerate(col_widths, 1):
                ws_summary.column_dimensions[chr(64 + i)].width = width
            
            # Sheet2-N: 个股详细报告
            for symbol in self.stock_pool:
                ws_report = wb.create_sheet(title=symbol[:10])
                report = self.generate_report(symbol)
                for line in report.split('\n'):
                    ws_report.append([line])
                ws_report.column_dimensions['A'].width = 70
            
            wb.save(filename)
            print(f"\n✅ 已导出到: {filename}")
            return filename
            
        except Exception as e:
            print(f"\n❌ 导出失败: {e}")
            return None


def demo():
    """
    演示函数：分析几只典型股票
    """
    print("\n" + "=" * 70)
    print("📊 股票分析器 V2.0 - Baostock版本")
    print("=" * 70)
    
    # 示例股票列表（可以修改为任意股票）
    demo_stocks = [
        'sh.600519',  # 贵州茅台
        'sh.601318',  # 中国平安
        'sz.000858',  # 五粮液
        'sh.600036',  # 招商银行
        'sz.300750',  # 宁德时代
        'sh.688981',  # 中芯国际
        'sz.002475',  # 立讯精密
        'sh.600276',  # 恒瑞医药
    ]
    
    # 创建分析器
    with StockAnalyzer() as analyzer:
        # 分析股票
        result = analyzer.analyze_stocks(demo_stocks)
        
        if result is not None:
            # 打印所有详细报告
            print("\n" + "=" * 70)
            print("📋 详细分析报告")
            print("=" * 70)
            analyzer.print_all_reports()
            
            # 导出Excel
            analyzer.export_to_excel()
    
    print("\n" + "=" * 70)
    print("✅ 分析完成！")
    print("=" * 70)


def custom_analysis(stock_list):
    """
    自定义分析函数
    参数: stock_list - 股票代码列表
    """
    print("\n" + "=" * 70)
    print("📊 自定义股票分析")
    print("=" * 70)
    
    with StockAnalyzer() as analyzer:
        result = analyzer.analyze_stocks(stock_list)
        
        if result is not None:
            analyzer.print_all_reports()
            analyzer.export_to_excel()
    
    print("\n" + "=" * 70)
    print("✅ 分析完成！")
    print("=" * 70)


if __name__ == '__main__':
    # 运行演示
    demo()