# coding=utf-8
from __future__ import print_function, absolute_import, unicode_literals
from gm.api import *
import datetime
import numpy as np
import pandas as pd
import logging
import json
import math

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler('stock_selector.log', encoding='utf-8'), logging.StreamHandler()]
)
logger = logging.getLogger(__name__)


class StockSelector:
    """
    股票池选股器
    基于东方财富掘金量化平台
    功能：多因子选股 + 个股分析报告
    输出：股票池 + Excel分析报告
    """

    def __init__(self, config_file='stock_selector_config.json'):
        self.load_config(config_file)
        self.stock_pool = []
        self.stock_analysis = {}

    def load_config(self, config_file):
        """加载配置文件"""
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
            logger.info(f'配置加载成功')
        except Exception as e:
            logger.error(f'配置加载失败：{e}')
            self.config = self._get_default_config()

    def _get_default_config(self):
        """获取默认配置"""
        return {
            'strategy': {
                'name': '多因子选股策略',
                'lookback_days_short': 10,
                'lookback_days_mid': 20,
                'lookback_days_long': 60,
                'holding_num': 20,
                'min_market_cap': 50,  # 最小市值（亿元）
                'max_pe': 100,
                'max_pb': 10
            },
            'indices': {
                'SHSE.000300': '沪深300',
                'SHSE.000905': '中证500',
                'SHSE.000016': '上证50'
            },
            'factors': {
                'momentum_short_weight': 0.15,
                'momentum_mid_weight': 0.20,
                'momentum_long_weight': 0.10,
                'valuation_pe_weight': 0.12,
                'valuation_pb_weight': 0.08,
                'valuation_ps_weight': 0.05,
                'volatility_weight': 0.10,
                'quality_roe_weight': 0.10,
                'growth_weight': 0.10
            },
            'backtest': {
                'token': '{{token}}'
            }
        }

    def judge_market_environment(self, date_str):
        """判断市场环境"""
        try:
            data = history_n(
                symbol='SHSE.000300',
                frequency='1d',
                count=20,
                fields='close',
                fill_missing='Last',
                adjust=ADJUST_PREV,
                end_time=date_str,
                df=True
            )

            if data is None or len(data) < 10:
                return 'neutral'

            returns = data['close'].pct_change().dropna()
            avg_return = returns.mean()
            volatility = returns.std()

            if avg_return > 0.002:
                return 'bull'
            elif avg_return < -0.002:
                return 'bear'
            elif volatility > 0.025:
                return 'volatile'
            else:
                return 'neutral'
        except Exception as e:
            logger.error(f'判断市场环境失败：{e}')
            return 'neutral'

    def _calculate_atr(self, high, low, close, period=14):
        """计算ATR"""
        tr1 = high[1:] - low[1:]
        tr2 = np.abs(high[1:] - close[:-1])
        tr3 = np.abs(low[1:] - close[:-1])
        tr = np.maximum(np.maximum(tr1, tr2), tr3)
        atr = np.zeros(len(close))
        if len(tr) >= period:
            atr[period] = np.mean(tr[:period])
            for i in range(period + 1, len(close)):
                atr[i] = (atr[i-1] * (period - 1) + tr[i-1]) / period
        return atr

    def _calculate_rsi(self, prices, period=14):
        """计算RSI"""
        deltas = np.diff(prices)
        gains = deltas.copy()
        gains[gains < 0] = 0
        losses = -deltas.copy()
        losses[losses < 0] = 0
        avg_gain = np.zeros(len(prices))
        avg_loss = np.zeros(len(prices))
        if len(gains) >= period:
            avg_gain[period] = np.mean(gains[:period])
            avg_loss[period] = np.mean(losses[:period])
            for i in range(period + 1, len(prices)):
                avg_gain[i] = (avg_gain[i-1] * (period - 1) + gains[i-1]) / period
                avg_loss[i] = (avg_loss[i-1] * (period - 1) + losses[i-1]) / period
        rs = avg_gain / (avg_loss + 1e-10)
        return 100 - (100 / (1 + rs))

    def _calculate_max_drawdown(self, prices):
        """计算最大回撤"""
        peak = prices[0]
        max_dd = 0
        for price in prices:
            if price > peak:
                peak = price
            dd = (price - peak) / peak
            if dd < max_dd:
                max_dd = dd
        return max_dd

    def _normalize(self, series):
        """Min-Max标准化"""
        min_val = series.min()
        max_val = series.max()
        return (series - min_val) / (max_val - min_val + 1e-10)

    def select_best_sector(self, end_date):
        """选择最佳风格板块"""
        market_env = self.judge_market_environment(end_date)
        logger.info(f'市场环境：{market_env}')

        scores = pd.DataFrame()
        weights = self.config['factors']

        for index_symbol, index_name in self.config['indices'].items():
            try:
                data = history_n(
                    symbol=index_symbol,
                    frequency='1d',
                    count=60,
                    fields='close,volume,high,low',
                    fill_missing='Last',
                    adjust=ADJUST_PREV,
                    end_time=end_date,
                    df=True
                )

                if data is None or len(data) < 20:
                    continue

                prices = data['close'].values
                volumes = data['volume'].values

                momentum_short = (prices[-1] / prices[-11]) - 1
                momentum_mid = (prices[-1] / prices[-21]) - 1
                momentum_long = (prices[-1] / prices[-61]) - 1

                returns = np.log(prices[1:] / prices[:-1])
                volatility = returns.std() * np.sqrt(252)

                rsi = self._calculate_rsi(prices)[-1]
                volume_rank = np.mean(volumes[-10:] > np.roll(volumes, 10)[-10:])

                score = (
                    momentum_short * weights.get('momentum_short_weight', 0.15) +
                    momentum_mid * weights.get('momentum_mid_weight', 0.20) +
                    momentum_long * weights.get('momentum_long_weight', 0.10) +
                    (1 - rsi / 100) * weights.get('valuation_pe_weight', 0.12) +
                    (-volatility) * weights.get('volatility_weight', 0.10) +
                    volume_rank * 0.1
                )

                scores.loc[index_symbol, 'score'] = score
                scores.loc[index_symbol, 'name'] = index_name
                scores.loc[index_symbol, 'momentum'] = momentum_mid
                scores.loc[index_symbol, 'volatility'] = volatility
                scores.loc[index_symbol, 'rsi'] = rsi

            except Exception as e:
                logger.error(f'计算指数因子失败 {index_symbol}: {e}')

        if scores.empty:
            return 'SHSE.000300'

        best_sector = scores['score'].idxmax()
        logger.info(f'最佳板块：{best_sector}({scores.loc[best_sector, "name"]})')
        logger.info(f'各指数评分:\n{scores[["name", "score", "momentum", "volatility", "rsi"]].round(4)}')

        return best_sector

    def get_stock_data(self, symbols, end_date):
        """获取股票多维度数据"""
        result = []
        total = len(symbols)
        logger.info(f'开始获取{total}只股票数据...')

        for i, symbol in enumerate(symbols):
            try:
                price_data = history_n(
                    symbol=symbol,
                    frequency='1d',
                    count=60,
                    fields='close,volume,high,low,amount',
                    fill_missing='Last',
                    adjust=ADJUST_PREV,
                    end_time=end_date,
                    df=True
                )

                if price_data is None or len(price_data) < 20:
                    continue

                prices = price_data['close'].values
                volumes = price_data['volume'].values
                high = price_data['high'].values
                low = price_data['low'].values
                amounts = price_data['amount'].values

                try:
                    fin_data = stk_get_financial_report(
                        symbol=symbol,
                        fields='pe_ttm,pb,ps,roe,eps,ebitda,revenue,profit',
                        report_type=REPORT_TYPE_QUARTER,
                        end_time=end_date,
                        df=True
                    )

                    pe_ttm = float(fin_data['pe_ttm'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['pe_ttm'].iloc[0]) else 30
                    pb = float(fin_data['pb'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['pb'].iloc[0]) else 3
                    ps = float(fin_data['ps'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['ps'].iloc[0]) else 2
                    roe = float(fin_data['roe'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['roe'].iloc[0]) else 0.1
                    eps = float(fin_data['eps'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['eps'].iloc[0]) else 0.1
                    revenue = float(fin_data['revenue'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['revenue'].iloc[0]) else 0
                    profit = float(fin_data['profit'].iloc[0]) if not fin_data.empty and not pd.isna(fin_data['profit'].iloc[0]) else 0
                except Exception as e:
                    logger.debug(f'获取财务数据失败 {symbol}: {e}')
                    pe_ttm, pb, ps, roe, eps, revenue, profit = 30, 3, 2, 0.1, 0.1, 0, 0

                try:
                    mv_data = stk_get_daily_mktvalue_pt(
                        symbols=[symbol],
                        fields='tot_mv,float_mv',
                        trade_date=end_date,
                        df=True
                    )
                    tot_mv = float(mv_data['tot_mv'].iloc[0]) if not mv_data.empty else 0
                    float_mv = float(mv_data['float_mv'].iloc[0]) if not mv_data.empty else 0
                except Exception as e:
                    logger.debug(f'获取市值数据失败 {symbol}: {e}')
                    tot_mv, float_mv = 0, 0

                short_days = self.config['strategy']['lookback_days_short']
                mid_days = self.config['strategy']['lookback_days_mid']
                long_days = self.config['strategy']['lookback_days_long']

                momentum_short = (prices[-1] / prices[-short_days-1]) - 1 if len(prices) > short_days else 0
                momentum_mid = (prices[-1] / prices[-mid_days-1]) - 1 if len(prices) > mid_days else 0
                momentum_long = (prices[-1] / prices[-long_days-1]) - 1 if len(prices) > long_days else 0

                returns = np.log(prices[1:] / prices[:-1])
                volatility = returns.std() * np.sqrt(252)

                atr = self._calculate_atr(high, low, prices)
                avg_atr = np.mean(atr[-20:]) / prices[-1] if len(atr) >= 20 else 0.02

                volume_trend = (volumes[-10:].mean() / volumes[-60:].mean()) - 1 if volumes[-60:].mean() > 0 else 0

                rsi = self._calculate_rsi(prices)[-1] if len(prices) >= 14 else 50

                max_dd = self._calculate_max_drawdown(prices)

                turnover_rate = volumes[-1] / (tot_mv / prices[-1] * 1e8) if tot_mv > 0 else 0

                profit_growth = (profit / (revenue + 1)) * 100

                result.append({
                    'symbol': symbol,
                    'name': symbol.split('.')[-1] if '.' in symbol else symbol,
                    'close': prices[-1],
                    'tot_mv': tot_mv,
                    'float_mv': float_mv,
                    'pe_ttm': pe_ttm,
                    'pb': pb,
                    'ps': ps,
                    'roe': roe,
                    'eps': eps,
                    'momentum_short': momentum_short,
                    'momentum_mid': momentum_mid,
                    'momentum_long': momentum_long,
                    'volatility': volatility,
                    'atr': avg_atr,
                    'volume_trend': volume_trend,
                    'rsi': rsi,
                    'max_drawdown': max_dd,
                    'turnover_rate': turnover_rate,
                    'profit_growth': profit_growth
                })

                if (i + 1) % 20 == 0:
                    logger.info(f'已处理 {i + 1}/{total} 只股票')

            except Exception as e:
                logger.error(f'处理股票 {symbol} 失败：{e}')

        return pd.DataFrame(result)

    def factor_scoring(self, stock_df):
        """多因子评分"""
        weights = self.config['factors']

        stock_df['momentum_short_score'] = self._normalize(stock_df['momentum_short'])
        stock_df['momentum_mid_score'] = self._normalize(stock_df['momentum_mid'])
        stock_df['momentum_long_score'] = self._normalize(stock_df['momentum_long'])

        stock_df['value_pe_score'] = 1 - self._normalize(stock_df['pe_ttm'].clip(upper=100))
        stock_df['value_pb_score'] = 1 - self._normalize(stock_df['pb'].clip(upper=10))
        stock_df['value_ps_score'] = 1 - self._normalize(stock_df['ps'].clip(upper=20))

        stock_df['volatility_score'] = 1 - self._normalize(stock_df['atr'])
        stock_df['quality_score'] = self._normalize(stock_df['roe'].clip(lower=0))
        stock_df['growth_score'] = self._normalize(stock_df['profit_growth'].clip(lower=-100, upper=100))

        stock_df['综合得分'] = (
            stock_df['momentum_short_score'] * weights.get('momentum_short_weight', 0.15) +
            stock_df['momentum_mid_score'] * weights.get('momentum_mid_weight', 0.20) +
            stock_df['momentum_long_score'] * weights.get('momentum_long_weight', 0.10) +
            stock_df['value_pe_score'] * weights.get('valuation_pe_weight', 0.12) +
            stock_df['value_pb_score'] * weights.get('valuation_pb_weight', 0.08) +
            stock_df['value_ps_score'] * weights.get('valuation_ps_weight', 0.05) +
            stock_df['volatility_score'] * weights.get('volatility_weight', 0.10) +
            stock_df['quality_score'] * weights.get('quality_roe_weight', 0.10) +
            stock_df['growth_score'] * weights.get('growth_weight', 0.10)
        )

        return stock_df

    def filter_stocks(self, stock_df):
        """基本面筛选"""
        min_mv = self.config['strategy'].get('min_market_cap', 50)
        max_pe = self.config['strategy'].get('max_pe', 100)
        max_pb = self.config['strategy'].get('max_pb', 10)

        filtered = stock_df[
            (stock_df['tot_mv'] >= min_mv) &
            (stock_df['pe_ttm'] > 0) &
            (stock_df['pe_ttm'] <= max_pe) &
            (stock_df['pb'] <= max_pb) &
            (stock_df['roe'] > 0)
        ].copy()

        logger.info(f'基本面筛选后剩余 {len(filtered)} 只股票')

        return filtered

    def select_stocks(self, end_date):
        """执行选股"""
        logger.info('=' * 60)
        logger.info('开始选股流程')
        logger.info('=' * 60)

        best_sector = self.select_best_sector(end_date)

        try:
            constituents = stk_get_index_constituents(
                index=best_sector,
                trade_date=end_date
            )

            if constituents is None or len(constituents) == 0:
                logger.error('获取成分股失败，尝试使用备选指数')
                best_sector = 'SHSE.000300'
                constituents = stk_get_index_constituents(
                    index=best_sector,
                    trade_date=end_date
                )

            symbols = list(constituents['symbol'])
            logger.info(f'获取{self.config["indices"].get(best_sector, best_sector)}成分股：{len(symbols)}只')

        except Exception as e:
            logger.error(f'获取成分股失败：{e}')
            return pd.DataFrame()

        stock_df = self.get_stock_data(symbols, end_date)

        if stock_df.empty:
            logger.error('未获取到任何股票数据')
            return pd.DataFrame()

        logger.info(f'获取数据完成，共 {len(stock_df)} 只股票')

        filtered_df = self.filter_stocks(stock_df)

        scored_df = self.factor_scoring(filtered_df)

        selected = scored_df.nlargest(
            self.config['strategy'].get('holding_num', 20),
            '综合得分'
        )

        self.stock_pool = list(selected['symbol'])

        for _, row in selected.iterrows():
            self.stock_analysis[row['symbol']] = {
                'symbol': row['symbol'],
                'name': row['name'],
                'score': row['综合得分'],
                'close': row['close'],
                'momentum_short': row['momentum_short'],
                'momentum_mid': row['momentum_mid'],
                'momentum_long': row['momentum_long'],
                'pe_ttm': row['pe_ttm'],
                'pb': row['pb'],
                'ps': row['ps'],
                'roe': row['roe'],
                'eps': row['eps'],
                'volatility': row['volatility'],
                'atr': row['atr'],
                'volume_trend': row['volume_trend'],
                'rsi': row['rsi'],
                'max_drawdown': row['max_drawdown'],
                'tot_mv': row['tot_mv'],
                'float_mv': row['float_mv'],
                'profit_growth': row['profit_growth']
            }

        logger.info('=' * 60)
        logger.info(f'选股完成，股票池共 {len(selected)} 只')
        logger.info(f'股票池：{self.stock_pool}')
        logger.info('=' * 60)

        return selected

    def generate_analysis_report(self, symbol):
        """生成个股分析报告"""
        if symbol not in self.stock_analysis:
            return None

        data = self.stock_analysis[symbol]

        report = []
        report.append('=' * 60)
        report.append(f'{data["name"]} ({data["symbol"]}) 个股分析报告')
        report.append('=' * 60)
        report.append('')

        report.append('【一、基本信息】')
        report.append(f'  股票代码：{data["symbol"]}')
        report.append(f'  最新收盘价：{data["close"]:.2f} 元')
        report.append(f'  总市值：{data["tot_mv"]:.2f} 亿元')
        report.append(f'  流通市值：{data["float_mv"]:.2f} 亿元')
        report.append('')

        report.append('【二、估值指标】')
        report.append(f'  市盈率(PE_TTM)：{data["pe_ttm"]:.2f}')
        report.append(f'  市净率(PB)：{data["pb"]:.2f}')
        report.append(f'  市销率(PS)：{data["ps"]:.2f}')
        pe_status = '偏低' if data['pe_ttm'] < 15 else ('合理' if data['pe_ttm'] < 30 else '偏高')
        report.append(f'  估值评价：{pe_status}')
        report.append('')

        report.append('【三、质量指标】')
        report.append(f'  净资产收益率(ROE)：{data["roe"]*100:.2f}%')
        report.append(f'  每股收益(EPS)：{data["eps"]:.4f} 元')
        report.append(f'  净利润增长率：{data["profit_growth"]:.2f}%')
        roe_status = '优秀' if data['roe'] > 0.15 else ('良好' if data['roe'] > 0.08 else '一般')
        report.append(f'  质量评价：{roe_status}')
        report.append('')

        report.append('【四、动量指标】')
        report.append(f'  短期动量(10日)：{data["momentum_short"]*100:.2f}%')
        report.append(f'  中期动量(20日)：{data["momentum_mid"]*100:.2f}%')
        report.append(f'  长期动量(60日)：{data["momentum_long"]*100:.2f}%')

        if data['momentum_mid'] > 0.1:
            momentum_status = '强势上涨'
        elif data['momentum_mid'] > 0:
            momentum_status = '温和上涨'
        elif data['momentum_mid'] > -0.1:
            momentum_status = '小幅回调'
        else:
            momentum_status = '明显下跌'
        report.append(f'  动量评价：{momentum_status}')
        report.append('')

        report.append('【五、风险指标】')
        report.append(f'  年化波动率：{data["volatility"]*100:.2f}%')
        report.append(f'  ATR指标：{data["atr"]*100:.2f}%')
        report.append(f'  历史最大回撤：{data["max_drawdown"]*100:.2f}%')
        volatility_status = '低波动' if data['volatility'] < 0.2 else ('中等波动' if data['volatility'] < 0.35 else '高波动')
        report.append(f'  风险评价：{volatility_status}')
        report.append('')

        report.append('【六、技术指标】')
        report.append(f'  RSI(14日)：{data["rsi"]:.2f}')
        report.append(f'  成交量趋势：{data["volume_trend"]*100:.2f}%')
        if data['rsi'] > 70:
            rsi_status = '超买区域'
        elif data['rsi'] < 30:
            rsi_status = '超卖区域'
        else:
            rsi_status = '正常区间'
        report.append(f'  RSI评价：{rsi_status}')
        report.append('')

        report.append('【七、综合评分】')
        report.append(f'  综合得分：{data["score"]:.4f}')
        if data['score'] > 0.7:
            overall = '强烈推荐'
        elif data['score'] > 0.5:
            overall = '推荐'
        elif data['score'] > 0.3:
            overall = '谨慎推荐'
        else:
            overall = '不推荐'
        report.append(f'  综合评价：{overall}')
        report.append('')

        report.append('=' * 60)

        return '\n'.join(report)

    def export_to_excel(self):
        """导出选股结果到Excel"""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.chart import RadarChart, Reference

            wb = openpyxl.Workbook()

            ws_pool = wb.active
            ws_pool.title = '股票池'

            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            ws_pool.append(['股票池汇总'])
            ws_pool['A1'].font = Font(bold=True, size=14)
            ws_pool.merge_cells('A1:R1')
            ws_pool.append([])

            ws_pool.append(['选股日期', datetime.datetime.now().strftime('%Y-%m-%d')])
            ws_pool.append(['股票数量', len(self.stock_pool)])
            ws_pool.append(['选股策略', self.config['strategy']['name']])
            ws_pool.append([])

            headers = ['排名', '股票代码', '股票名称', '收盘价', '综合得分', 'PE', 'PB', 'ROE(%)',
                      '短期动量(%)', '中期动量(%)', '长期动量(%)', '波动率(%)', 'RSI',
                      '总市值(亿)', '净利润增长(%)', 'ATR(%)', '成交量趋势(%)', '最大回撤(%)']

            ws_pool.append(headers)
            for cell in ws_pool[ws_pool.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            sorted_stocks = sorted(
                self.stock_analysis.values(),
                key=lambda x: x['score'],
                reverse=True
            )

            for rank, stock in enumerate(sorted_stocks, 1):
                row_data = [
                    rank,
                    stock['symbol'],
                    stock['name'],
                    round(stock['close'], 2),
                    round(stock['score'], 4),
                    round(stock['pe_ttm'], 2),
                    round(stock['pb'], 2),
                    round(stock['roe'] * 100, 2),
                    round(stock['momentum_short'] * 100, 2),
                    round(stock['momentum_mid'] * 100, 2),
                    round(stock['momentum_long'] * 100, 2),
                    round(stock['volatility'] * 100, 2),
                    round(stock['rsi'], 2),
                    round(stock['tot_mv'], 2),
                    round(stock['profit_growth'], 2),
                    round(stock['atr'] * 100, 2),
                    round(stock['volume_trend'] * 100, 2),
                    round(stock['max_drawdown'] * 100, 2)
                ]
                ws_pool.append(row_data)

            for row in ws_pool.iter_rows(min_row=ws_pool.min_row + 6, max_row=ws_pool.max_row):
                for cell in row:
                    cell.border = thin_border
                    if cell.column > 1:
                        cell.alignment = Alignment(horizontal='right')
                    if cell.row > ws_pool.min_row + 5:
                        if cell.column == 5:
                            if cell.value > 0.6:
                                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            elif cell.value > 0.4:
                                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                            else:
                                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            col_widths = [6, 15, 12, 10, 10, 8, 8, 10, 12, 12, 12, 10, 8, 12, 12, 10, 12, 12]
            for i, width in enumerate(col_widths, 1):
                ws_pool.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            ws_pool.row_dimensions[1].height = 25

            for i, stock in enumerate(sorted_stocks, 2):
                ws_report = wb.create_sheet(title=f'{stock["name"][:8]}')

                report = self.generate_analysis_report(stock['symbol'])
                lines = report.split('\n')

                title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

                for line in lines:
                    if '=' in line and line.strip().startswith('='):
                        ws_report.append([line])
                        ws_report[ws_report.max_row][0].fill = title_fill
                        ws_report[ws_report.max_row][0].font = Font(bold=True, size=12, color='FFFFFF')
                        ws_report.merge_cells(f'A{ws_report.max_row}:D{ws_report.max_row}')
                    elif '【' in line and '】' in line:
                        ws_report.append([line])
                        ws_report[ws_report.max_row][0].font = Font(bold=True, size=11)
                        ws_report[ws_report.max_row][0].fill = section_fill
                    elif '：' in line:
                        ws_report.append([line])
                    elif line.strip():
                        ws_report.append([line])

                ws_report.column_dimensions['A'].width = 80

            ws_config = wb.create_sheet(title='策略配置')
            ws_config.append(['配置项', '值'])
            ws_config['A1'].font = Font(bold=True)
            ws_config['B1'].font = Font(bold=True)

            self._flatten_config(self.config, '', ws_config)
            ws_config.column_dimensions['A'].width = 30
            ws_config.column_dimensions['B'].width = 50

            filename = f'股票池_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(filename)
            logger.info(f'选股结果已导出到: {filename}')

            return filename

        except ImportError:
            logger.error('需要安装 openpyxl 库: pip install openpyxl')
        except Exception as e:
            logger.error(f'导出Excel失败：{e}')

    def _flatten_config(self, config, prefix, ws):
        """扁平化配置"""
        for key, value in config.items():
            new_key = f'{prefix}.{key}' if prefix else key
            if isinstance(value, dict):
                self._flatten_config(value, new_key, ws)
            else:
                ws.append([new_key, value])

    def print_summary(self):
        """打印选股汇总"""
        logger.info('=' * 80)
        logger.info('选股结果汇总')
        logger.info('=' * 80)

        sorted_stocks = sorted(
            self.stock_analysis.values(),
            key=lambda x: x['score'],
            reverse=True
        )

        for rank, stock in enumerate(sorted_stocks, 1):
            logger.info(f'{rank:2d}. {stock["symbol"]} | 得分:{stock["score"]:.4f} | '
                        f'PE:{stock["pe_ttm"]:.1f} | ROE:{stock["roe"]*100:.1f}% | '
                        f'动量:{stock["momentum_mid"]*100:+.1f}% | '
                        f'RSI:{stock["rsi"]:.0f} | '
                        f'市值:{stock["tot_mv"]:.0f}亿')

            report = self.generate_analysis_report(stock['symbol'])
            if report:
                print(report)
                print()

        logger.info('=' * 80)
        logger.info(f'共选出 {len(self.stock_pool)} 只股票')
        logger.info('=' * 80)


selector = StockSelector()


def init(context):
    """初始化"""
    logger.info('股票池选股器初始化完成')
    logger.info(f'选股策略: {selector.config["strategy"]["name"]}')


def on_bar(context, bars):
    """K线数据回调（用于实时选股）"""
    for bar in bars:
        logger.info(f'{bar.symbol} 最新价: {bar.close}')


def on_schedule(context, bar_list):
    """定时任务回调"""
    now_str = context.now.strftime('%Y-%m-%d')

    logger.info(f'=' * 60)
    logger.info(f'开始执行选股任务 {now_str}')
    logger.info(f'=' * 60)

    selected = selector.select_stocks(now_str)

    if not selected.empty:
        selector.print_summary()
        selector.export_to_excel()
    else:
        logger.warning('选股结果为空')


def on_backtest_finished(context, indicator):
    """回测结束回调"""
    logger.info('选股任务完成')


if __name__ == '__main__':
    """运行选股器"""
    import sys

    mode = 'schedule' if len(sys.argv) > 1 and sys.argv[1] == '--schedule' else 'standalone'

    now_date = datetime.datetime.now().strftime('%Y-%m-%d')

    logger.info('=' * 60)
    logger.info('股票池选股器启动')
    logger.info(f'执行日期: {now_date}')
    logger.info('=' * 60)

    selected = selector.select_stocks(now_date)

    if not selected.empty:
        selector.print_summary()
        filename = selector.export_to_excel()
        logger.info(f'Excel文件已生成: {filename}')
    else:
        logger.warning('选股结果为空')

    if mode == 'schedule':
        run(
            strategy_id='stock_selector',
            filename='stock_selector.py',
            mode=MODE_LIVE,
            token=selector.config['backtest']['token'],
            schedule_market='SHSE',
            schedule_func=on_schedule,
            running_mode=RUNING_MODE_ONLINE
        )